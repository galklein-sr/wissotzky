import re
from typing import List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from Logic.w74_helper_suppression import suppress_rows_by_helper

SUM_ANCHOR_AFTER_TODAY = 'סה"כ סכום יתרת חוב עד היום'
SUM_ANCHOR_TOTAL = 'סה"כ סכום יתרת חוב'
HELPER_COL_NAME = "טור עזר"

FORBIDDEN = r'[:\\/?*\[\]]'

def _sanitize_sheet_name(name: str, used: set) -> str:
    s = re.sub(FORBIDDEN, " ", str(name)).strip()
    if not s:
        s = "Sheet"
    s = s[:31]  # Excel limit
    base = s
    i = 1
    while s in used:
        suffix = f"_{i}"
        s = (base[:31-len(suffix)] + suffix) if len(base)+len(suffix) > 31 else base + suffix
        i += 1
    used.add(s)
    return s

def _sanitize_sheet_name_preview(name: str) -> str:
    """סניטציה ללא רישום ל-used: שימוש לזיהוי לשוניות מנהלים קיימות למחיקה עדינה."""
    s = re.sub(FORBIDDEN, " ", str(name)).strip()
    if not s:
        s = "Sheet"
    return s[:31]

def _pick_first(df_cols: List[str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df_cols:
            return c
    return None

def _dynamic_month_cols(df_cols: List[str], anchor: str, max_cols: int = 4) -> List[str]:
    if anchor in df_cols:
        idx = df_cols.index(anchor)
        return [c for c in df_cols[idx+1 : idx+1+max_cols] if c in df_cols]
    return []

def _autosize(ws):
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def _style_header(ws):
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for j in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

def _add_column_sums_row(ws, header_names: List[str], total_col: str, today_col: str):
    """
    מוסיף שורת סכומים לכל העמודות מ-H עד N (כולל) שתי שורות מתחת לשורה האחרונה.
    מניח שהעמודות כתובות לפי הסדר: ... G('קוד סוכן'), H(total_col), I(today_col), J..M(dyn), N(HELPER_COL_NAME)
    """
    if total_col not in header_names or today_col not in header_names or HELPER_COL_NAME not in header_names:
        return  # אין מבנה מלא, לא כותבים סיכום

    name_to_idx0 = {name: i for i, name in enumerate(header_names)}
    first_sum_idx0 = name_to_idx0[total_col]       # H (0-based)
    last_sum_idx0  = name_to_idx0[HELPER_COL_NAME] # N (0-based)
    if last_sum_idx0 < first_sum_idx0:
        return

    last_row = ws.max_row
    sum_row = last_row + 2  # שתי שורות מתחת לשורה האחרונה

    # "סכום :" בעמודה שלפני H -> כלומר G (עמודה מספר first_sum_idx0 ב-1-based)
    label_col_excel = first_sum_idx0
    if label_col_excel < 1:
        label_col_excel = 1
    ws.cell(row=sum_row, column=label_col_excel).value = "סכום :"
    ws.cell(row=sum_row, column=label_col_excel).font = Font(bold=True)

    start_data_row = 2  # אחרי שורת כותרת
    for idx0 in range(first_sum_idx0, last_sum_idx0 + 1):
        col_excel = idx0 + 1
        col_letter = get_column_letter(col_excel)
        ws.cell(row=sum_row, column=col_excel).value = f"=SUM({col_letter}{start_data_row}:{col_letter}{last_row})"
        ws.cell(row=sum_row, column=col_excel).number_format = '#,##0.00'
        ws.cell(row=sum_row, column=col_excel).font = Font(bold=True)

def build_manager_sheets(
    processed_df: pd.DataFrame,
    output_path: str,
    managers_col: str = "מנהל סחר",
    max_month_cols_after_today: int = 4
) -> Tuple[int, List[str]]:
    """
    יוצר לשונית לכל 'מנהל סחר' מתוך DF של 'מעובד', בתוך אותו קובץ אקסל (output_path).
    מבנה עמודות:
    A מנהל סחר | B מנהל אזור | C סוכן | D ערוץ | E קוד לקוח משלם | F לקוח משלם |
    G קוד סוכן | H סכום יתרת חוב | I סכום יתרת חוב עד היום | J..M חודשים דינמיים אחרי I (עד 4) | N טור עזר = sum(J..M)
    ובסוף: שורת סיכום (שתי שורות מתחת לשורה האחרונה) עם "סכום :" וסכומים לכל H..N.
    """
    df = processed_df.copy()
    cols = list(df.columns)
    if managers_col not in cols:
        return 0, []

    region_col = _pick_first(cols, ["מנהל אזור", "מנהל איזור"])
    agent_col = "סוכן" if "סוכן" in cols else None
    channel_col = "ערוץ" if "ערוץ" in cols else None
    agent_code_col = "קוד סוכן" if "קוד סוכן" in cols else None

    total_col = SUM_ANCHOR_TOTAL if SUM_ANCHOR_TOTAL in cols else None
    today_col = SUM_ANCHOR_AFTER_TODAY if SUM_ANCHOR_AFTER_TODAY in cols else None

    payer_code_src = _pick_first(cols, ["קוד לקוח קצה", "קוד לקוח משלם"])
    payer_name_src = _pick_first(cols, ["לקוח קצה", "לקוח משלם"])

    dyn_cols = _dynamic_month_cols(cols, SUM_ANCHOR_AFTER_TODAY, max_month_cols_after_today)
    if not dyn_cols and total_col:
        dyn_cols = _dynamic_month_cols(cols, SUM_ANCHOR_TOTAL, max_month_cols_after_today)

    managers = (
        df[managers_col].dropna().astype(str).str.strip().replace({"": None}).dropna().unique().tolist()
    )
    managers = sorted(managers)

    wb = load_workbook(output_path)

    # --- מחיקה עדינה: רק לשוניות מנהלים קודמות ---
    existing = list(wb.sheetnames)
    expected_bases = set(_sanitize_sheet_name_preview(m) for m in managers)
    to_delete = []
    for sh in existing:
        if sh == "מעובד":
            continue
        # מחיקה אם השם תואם לשם מנהל מסונן או עם סיומת _1, _2...
        for base in expected_bases:
            if sh == base or sh.startswith(base + "_"):
                to_delete.append(sh)
                break
    for sh in to_delete:
        del wb[sh]
    # ----------------------------------------------

    used_names = set(wb.sheetnames)
    created = []

    for m in managers:
        sub = df.loc[df[managers_col].astype(str).str.strip() == m].copy()
        if sub.empty:
            continue

        col_order = []
        data = {}

        col_order += ["מנהל סחר", "מנהל אזור", "סוכן", "ערוץ"]
        data["מנהל סחר"] = sub[managers_col] if managers_col in sub.columns else ""
        data["מנהל אזור"] = sub[region_col] if region_col in sub.columns else ""
        data["סוכן"] = sub[agent_col] if agent_col in sub.columns else ""
        data["ערוץ"] = sub[channel_col] if channel_col in sub.columns else ""

        col_order += ["קוד לקוח משלם", "לקוח משלם"]
        data["קוד לקוח משלם"] = sub[payer_code_src] if payer_code_src in sub.columns else ""
        data["לקוח משלם"] = sub[payer_name_src] if payer_name_src in sub.columns else ""

        col_order += ["קוד סוכן"]
        data["קוד סוכן"] = sub[agent_code_col] if agent_code_col in sub.columns else ""

        if total_col:
            col_order += [total_col]
            data[total_col] = pd.to_numeric(sub[total_col], errors="coerce")
        else:
            col_order += [SUM_ANCHOR_TOTAL]
            data[SUM_ANCHOR_TOTAL] = ""

        if today_col:
            col_order += [today_col]
            data[today_col] = pd.to_numeric(sub[today_col], errors="coerce")
        else:
            col_order += [SUM_ANCHOR_AFTER_TODAY]
            data[SUM_ANCHOR_AFTER_TODAY] = ""

        for c in dyn_cols:
            col_order.append(c)
            data[c] = pd.to_numeric(sub[c], errors="coerce") if c in sub.columns else ""

        # N: טור עזר
        col_order.append(HELPER_COL_NAME)
        if dyn_cols:
            dyn_num = pd.concat([pd.to_numeric(sub[c], errors="coerce") for c in dyn_cols], axis=1)
            data[HELPER_COL_NAME] = dyn_num.sum(axis=1, skipna=True)
        else:
            data[HELPER_COL_NAME] = 0.0

        out_df = pd.DataFrame(data)[col_order]
        
        out_df = suppress_rows_by_helper(out_df, month_cols=dyn_cols, helper_col=HELPER_COL_NAME, threshold=-1000)


        sheet_name = _sanitize_sheet_name(m, used_names)
        ws = wb.create_sheet(title=sheet_name)

        ws.append(out_df.columns.tolist())
        for _, row in out_df.iterrows():
            ws.append(row.tolist())

        _style_header(ws)
        _add_column_sums_row(
            ws,
            header_names=out_df.columns.tolist(),
            total_col=(total_col or SUM_ANCHOR_TOTAL),
            today_col=(today_col or SUM_ANCHOR_AFTER_TODAY),
        )
        _autosize(ws)
        created.append(sheet_name)
        


    wb.save(output_path)
    return len(created), created
