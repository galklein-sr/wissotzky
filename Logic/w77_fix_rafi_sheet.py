import re
from typing import Tuple, List
from openpyxl import load_workbook

MANAGER_HEADER = "מנהל סחר"
REGION_HEADERS = ("מנהל אזור", "מנהל איזור")

def _match(value, pat: re.Pattern) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    return bool(pat.search(s))

def refine_rafi_sheet_rows(
    xlsx_path: str,
    target_base: str = "רפי מור יוסף",
    display_text: str = "רפי מור יוסף- סחר",
) -> Tuple[bool, List[str], List[int]]:
    """
    בכל לשונית ששמה מכיל 'רפי מור יוסף':
      1) מוחק שורות שבהן בעמודת 'מנהל אזור/איזור' אין התאמה ל'רפי מור יוסף' (כולל וריאציה '- סחר').
      2) מעדכן בשורות שנותרו את 'מנהל סחר' וגם את 'מנהל אזור/איזור' ל-display_text.
    """
    wb = load_workbook(xlsx_path)
    touched, deleted_counts = [], []

    # התאמה: 'רפי מור יוסף' עם/בלי ' - סחר'
    pat = re.compile(rf"{re.escape(target_base)}(\s*-\s*סחר)?")

    for sh in list(wb.sheetnames):
        if target_base not in sh:
            continue
        ws = wb[sh]

        # כותרות (שורה 1)
        headers = {str(c.value).strip(): i for i, c in enumerate(ws[1], start=1) if c.value}
        col_manager = headers.get(MANAGER_HEADER)
        col_region = next((headers[h] for h in REGION_HEADERS if h in headers), None)
        if not col_region:
            continue

        # סימון שורות למחיקה אם "מנהל אזור/איזור" לא רפי
        to_delete = []
        for r in range(2, ws.max_row + 1):
            if not _match(ws.cell(r, col_region).value, pat):
                to_delete.append(r)

        # מחיקה מלמטה למעלה
        for r in reversed(to_delete):
            ws.delete_rows(r, 1)

        # נרמול טקסטים לשורות שנותרו
        if col_manager:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_manager).value = display_text
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col_region).value = display_text

        if to_delete:
            touched.append(sh)
            deleted_counts.append(len(to_delete))

    if touched:
        wb.save(xlsx_path)
        wb.close()
        return True, touched, deleted_counts

    wb.close()
    return False, [], []