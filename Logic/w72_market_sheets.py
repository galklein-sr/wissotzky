from typing import List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re

from Logic.w74_helper_suppression import suppress_rows_by_helper

SUM_ANCHOR_AFTER_TODAY = 'סה"כ סכום יתרת חוב עד היום'  # I
SUM_ANCHOR_TOTAL      = 'סה"כ סכום יתרת חוב'          # H
HELPER_COL_NAME       = "טור עזר"



def _norm_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    # הסרת סימני RTL שקופים
    s = s.replace('\u200f','').replace('\u200e','')
    # איחוד מקפים שונים ל־"-"
    s = s.replace('–','-').replace('—','-').replace('-','-').replace('־','-')
    # רווחים אחידים סביב מקף
    s = re.sub(r'\s*-\s*', ' - ', s)
    # דחיסת רווחים
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

# ---------------- helpers ----------------
def _pick_first(df_cols: List[str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df_cols:
            return c
    return None

def _dynamic_month_cols(df_cols: List[str], anchor: str, max_cols: int = 4) -> List[str]:
    if anchor in df_cols:
        idx = df_cols.index(anchor)
        return [c for c in df_cols[idx+1: idx+1+max_cols] if c in df_cols]
    return []

def _style_header(ws):
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for j in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

def _autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def _add_column_sums_row(ws, header_names: List[str], total_col: str, today_col: str):
    # סכום לכל H..N, שתי שורות אחרי הנתונים
    name_to_idx0 = {n:i for i,n in enumerate(header_names)}
    if total_col not in name_to_idx0 or today_col not in name_to_idx0 or HELPER_COL_NAME not in name_to_idx0:
        return
    first_sum_idx0 = name_to_idx0[total_col]           # H
    last_sum_idx0  = name_to_idx0[HELPER_COL_NAME]     # N
    if last_sum_idx0 < first_sum_idx0:
        return

    last_row = ws.max_row
    sum_row  = last_row + 2
    label_col_excel = max(1, first_sum_idx0)           # G (עמודה לפני H)
    ws.cell(row=sum_row, column=label_col_excel).value = "סכום :"
    ws.cell(row=sum_row, column=label_col_excel).font  = Font(bold=True)

    start_data = 2
    for idx0 in range(first_sum_idx0, last_sum_idx0+1):
        col_excel  = idx0 + 1
        col_letter = get_column_letter(col_excel)
        c = ws.cell(row=sum_row, column=col_excel)
        c.value = f"=SUM({col_letter}{start_data}:{col_letter}{last_row})"
        c.number_format = '#,##0.00'
        c.font = Font(bold=True)

def _build_manager_like_df(sub: pd.DataFrame, max_month_cols_after_today: int = 4) -> Tuple[pd.DataFrame, str, List[str]]:
    """
    בונה DF במבנה זהה לגיליונות מנהל סחר:
    A מנהל סחר | B מנהל אזור | C סוכן | D ערוץ | E קוד לקוח משלם | F לקוח משלם |
    G קוד סוכן | H סה"כ סכום יתרת חוב | I סה"כ סכום יתרת חוב עד היום | J..M חודשי דינמי | N טור עזר
    מחזיר: (out_df, total_col_for_sum, month_cols_used)
    """
    cols = list(sub.columns)
    mgr_col     = "מנהל סחר" if "מנהל סחר" in cols else None
    region_col  = _pick_first(cols, ["מנהל אזור","מנהל איזור"])
    agent_col   = "סוכן" if "סוכן" in cols else None
    channel_col = "ערוץ" if "ערוץ" in cols else None
    payer_code  = _pick_first(cols, ["קוד לקוח משלם","קוד לקוח קצה"])
    payer_name  = _pick_first(cols, ["לקוח משלם","לקוח קצה"])
    agent_code  = "קוד סוכן" if "קוד סוכן" in cols else None

    total_col = SUM_ANCHOR_TOTAL if SUM_ANCHOR_TOTAL in cols else None
    today_col = SUM_ANCHOR_AFTER_TODAY if SUM_ANCHOR_AFTER_TODAY in cols else None

    # חודשי J..M — קודם ננסה לעגן ל-I, ואם לא קיים נעגן ל-H
    dyn = _dynamic_month_cols(cols, SUM_ANCHOR_AFTER_TODAY, max_month_cols_after_today)
    if not dyn and total_col:
        dyn = _dynamic_month_cols(cols, SUM_ANCHOR_TOTAL, max_month_cols_after_today)

    data = {}
    order: List[str] = []

    order += ["מנהל סחר","מנהל אזור","סוכן","ערוץ"]
    data["מנהל סחר"] = sub[mgr_col] if mgr_col in sub.columns else ""
    data["מנהל אזור"] = sub[region_col] if region_col in sub.columns else ""
    data["סוכן"] = sub[agent_col] if agent_col in sub.columns else ""
    data["ערוץ"] = sub[channel_col] if channel_col in sub.columns else ""

    order += ["קוד לקוח משלם","לקוח משלם"]
    data["קוד לקוח משלם"] = sub[payer_code] if payer_code in sub.columns else ""
    data["לקוח משלם"] = sub[payer_name] if payer_name in sub.columns else ""

    order += ["קוד סוכן"]
    data["קוד סוכן"] = sub[agent_code] if agent_code in sub.columns else ""

    if total_col:
        order += [total_col]
        data[total_col] = pd.to_numeric(sub[total_col], errors="coerce")
    else:
        order += [SUM_ANCHOR_TOTAL]
        data[SUM_ANCHOR_TOTAL] = ""

    if today_col:
        order += [today_col]
        data[today_col] = pd.to_numeric(sub[today_col], errors="coerce")
    else:
        order += [SUM_ANCHOR_AFTER_TODAY]
        data[SUM_ANCHOR_AFTER_TODAY] = ""

    for c in dyn:
        order.append(c)
        data[c] = pd.to_numeric(sub[c], errors="coerce") if c in sub.columns else None

    # N: טור עזר = סכום J..M
    order.append(HELPER_COL_NAME)
    if dyn:
        dyn_num = pd.concat([pd.to_numeric(sub[c], errors="coerce") for c in dyn], axis=1)
        data[HELPER_COL_NAME] = dyn_num.sum(axis=1, skipna=True)
    else:
        data[HELPER_COL_NAME] = 0.0

    out_df = pd.DataFrame(data)[order]
    total_for_sum = total_col or SUM_ANCHOR_TOTAL
    return out_df, total_for_sum, dyn

# ---------------- builders ----------------
def build_private_market_like_manager(
    df_processed: pd.DataFrame,
    output_path: str,
    manager_name: str = "רפי מור יוסף-סחר",
    channel_value: str = "שוק פרטי",
    max_month_cols_after_today: int = 4,
    sheet_name: str = "שוק פרטי"
) -> Tuple[bool, str]:
    """
    סינון לפי מנהל סחר + ערוץ, בניה במבנה מנהלים, דיכוי שורות לפי טור עזר < -1000, סכום H..N, כתיבה לגליון.
    """
    if df_processed.empty:
        return False, sheet_name

    cols = list(df_processed.columns)
    mgr_col = "מנהל סחר" if "מנהל סחר" in cols else None
    ch_col  = "ערוץ" if "ערוץ" in cols else None
    if not mgr_col or not ch_col:
        return False, sheet_name

    mgr_series_norm = df_processed[mgr_col].astype(str).map(_norm_text)
    ch_series_norm  = df_processed[ch_col].astype(str).map(_norm_text)
    target_mgr = _norm_text(manager_name)
    target_ch  = _norm_text(channel_value)

    mask = (mgr_series_norm == target_mgr) & (ch_series_norm == target_ch)
    sub = df_processed.loc[mask].copy()

    # לוג בקרה מעודכן:
    total_rows = len(df_processed)
    mgr_hits   = (mgr_series_norm == target_mgr).sum()
    ch_hits    = (ch_series_norm  == target_ch).sum()
    print(f"[שוק פרטי] סה\"כ שורות: {total_rows} | התאמות מנהל(נרמל): {mgr_hits} | התאמות ערוץ(נרמל): {ch_hits} | חיתוך: {len(sub)}", flush=True)

    # לבקרה: כמה שורות נמצאו
    print(f"[שוק פרטי] סה\"כ שורות: {len(df_processed)} | התאמות מנהל: {(df_processed[mgr_col].astype(str).str.strip()==manager_name).sum()} | התאמות ערוץ: {(df_processed[ch_col].astype(str).str.strip()==channel_value).sum()} | חיתוך: {len(sub)}", flush=True)

    if sub.empty:
        return False, sheet_name

    out_df, total_for_sum, dyn = _build_manager_like_df(sub, max_month_cols_after_today=max_month_cols_after_today)

    # דיכוי לפי טור עזר < -1000 על בסיס החודשים שבפועל יצאו לגליון
    cols_out = out_df.columns.tolist()
    # נזהה מחדש את חודשי ה-J..M לפי העוגן I (ואם לא, אז H) בתוך ה-DF הסופי:
    dyn_for_suppress = _dynamic_month_cols(cols_out, SUM_ANCHOR_AFTER_TODAY, max_month_cols_after_today)
    if not dyn_for_suppress:
        dyn_for_suppress = _dynamic_month_cols(cols_out, SUM_ANCHOR_TOTAL, max_month_cols_after_today)

    before = len(out_df)
    out_df2 = suppress_rows_by_helper(out_df, month_cols=dyn_for_suppress, helper_col=HELPER_COL_NAME, threshold=-1000)
    # כמה שורות סומנו (כלומר קיבלו '-' ו-J..M ריק):
    marked = (pd.to_numeric(out_df[HELPER_COL_NAME], errors="coerce") < -1000).sum()
    print(f"[שוק פרטי] דוכאו (helper<-1000): {marked} שורות", flush=True)
    out_df = out_df2

    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    ws.append(out_df.columns.tolist())
    for _, row in out_df.iterrows():
        ws.append(row.tolist())

    _style_header(ws)
    _add_column_sums_row(ws, header_names=out_df.columns.tolist(),
                         total_col=SUM_ANCHOR_TOTAL, today_col=SUM_ANCHOR_AFTER_TODAY)
    _autosize(ws)
    wb.save(output_path)
    return True, sheet_name

def build_tedmiti_full_columns(
    df_processed: pd.DataFrame,
    output_path: str,
    manager_name: str = "עמי חכמון",
    sheet_name: str = "שוק תדמיתי",
    max_month_cols_after_today: int = 4
) -> Tuple[bool, str]:
    """
    סינון לפי מנהל סחר בלבד, בניה במבנה מנהלים, דיכוי לפי טור עזר < -1000, סכום H..N, כתיבה לגליון.
    """
    if df_processed.empty:
        return False, sheet_name

    cols = list(df_processed.columns)
    mgr_col = "מנהל סחר" if "מנהל סחר" in cols else None
    if not mgr_col:
        return False, sheet_name

    sub = df_processed.loc[(df_processed[mgr_col].astype(str).str.strip() == manager_name)].copy()
    if sub.empty:
        return False, sheet_name

    out_df, total_for_sum, dyn = _build_manager_like_df(sub, max_month_cols_after_today=max_month_cols_after_today)

    cols_out = out_df.columns.tolist()
    dyn_for_suppress = _dynamic_month_cols(cols_out, SUM_ANCHOR_AFTER_TODAY, max_month_cols_after_today)
    if not dyn_for_suppress:
        dyn_for_suppress = _dynamic_month_cols(cols_out, SUM_ANCHOR_TOTAL, max_month_cols_after_today)

    out_df = suppress_rows_by_helper(out_df, month_cols=dyn_for_suppress, helper_col=HELPER_COL_NAME, threshold=-1000)

    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    ws.append(out_df.columns.tolist())
    for _, row in out_df.iterrows():
        ws.append(row.tolist())

    _style_header(ws)
    _add_column_sums_row(ws, header_names=out_df.columns.tolist(),
                         total_col=SUM_ANCHOR_TOTAL, today_col=SUM_ANCHOR_AFTER_TODAY)
    _autosize(ws)
    wb.save(output_path)
    return True, sheet_name