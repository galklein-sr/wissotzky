from __future__ import annotations
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from openpyxl.styles import Border, Side


AGENT_KEYS = ["תוויות שורה", "סוכן", "Row Labels"]
TOTAL_KEY  = "סה\"כ סכום יתרת חוב"
TODAY_KEY  = "סה\"כ סכום יתרת חוב עד היום"
MONTH_HINTS = ["חודש", "טרם"]

# ===== helpers to read pivots =====
def _row_values(ws: Worksheet, r: int) -> List[str]:
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        out.append("" if v is None else str(v).strip())
    return out

def _find_header_row(ws: Worksheet, scan_rows: int = 40) -> Optional[int]:
    # קודם מנסה לפי עמודה A
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        a = ws.cell(row=r, column=1).value
        if a and str(a).strip() in AGENT_KEYS:
            return r
    # אחרת מוצא "השורה הכי עשירה" שמכילה אחד הרמזים
    candidate = None
    best_non_empty = 0
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        vals = _row_values(ws, r)
        non_empty = sum(1 for x in vals if x)
        line = " | ".join(vals)
        if non_empty > best_non_empty and (
            any(k in line for k in AGENT_KEYS) or
            TOTAL_KEY in line or TODAY_KEY in line or
            any(h in line for h in MONTH_HINTS)
        ):
            best_non_empty = non_empty
            candidate = r
    return candidate

def _headers_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    return {
        ("" if ws.cell(row=header_row, column=c).value is None else str(ws.cell(row=header_row, column=c).value).strip()): c
        for c in range(1, ws.max_column + 1)
    }

def _first_existing(hmap: Dict[str, int], keys: List[str]) -> Optional[int]:
    for k in keys:
        if k in hmap:
            return hmap[k]
    return None

def _find_col_contains(hmap: Dict[str, int], must_contain: List[str]) -> Optional[int]:
    def norm(s: str) -> str:
        return s.replace("סכום של", "").replace("Σ", "").strip()
    for k, idx in hmap.items():
        if all(sub in norm(k) for sub in must_contain):
            return idx
    return None

def _collect_month_cols(hmap: Dict[str, int]) -> List[int]:
    pairs = []
    for k, idx in hmap.items():
        if any(h in k for h in MONTH_HINTS):
            pairs.append((idx, k))
    pairs.sort(key=lambda t: t[0])
    return [idx for idx, _ in pairs]

def _extract_from_pivot(ws: Worksheet, max_month_cols: int = 4) -> Tuple[List[Dict], List[str]]:
    header_row = _find_header_row(ws)
    if not header_row:
        return [], []
    hmap = _headers_map(ws, header_row)
    agent_col_idx = _first_existing(hmap, AGENT_KEYS) or 1
    total_col_idx = _find_col_contains(hmap, [TOTAL_KEY])
    today_col_idx = _find_col_contains(hmap, [TODAY_KEY])

    # fallback: אם לא זיהינו סכומים, קח 2 עמודות מספריות ראשונות אחרי סוכן
    if total_col_idx is None and today_col_idx is None:
        numeric_candidates = []
        for c in range(agent_col_idx + 1, ws.max_column + 1):
            v = ws.cell(row=header_row + 1, column=c).value
            if isinstance(v, (int, float)):
                numeric_candidates.append(c)
            if len(numeric_candidates) >= 2:
                break
        if numeric_candidates:
            total_col_idx = numeric_candidates[0]
            today_col_idx = numeric_candidates[1] if len(numeric_candidates) > 1 else None

    month_cols_all = _collect_month_cols(hmap)[:max_month_cols]
    # שמות הכותרות לחודשים
    month_headers = []
    for cidx in month_cols_all:
        for k, v in hmap.items():
            if v == cidx:
                month_headers.append(k); break

    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ws.max_column + 1)):
            break
        agent = ws.cell(row=r, column=agent_col_idx).value
        if agent and str(agent).strip() not in {"Grand Total", "סה\"כ", "סהכ", "Total", "סך הכל", "סכום כולל"}:
            def g(coli):
                return ws.cell(row=r, column=coli).value if coli else None
            rec = {
                "agent": str(agent).strip(),
                "F": g(total_col_idx) or 0,
                "G": g(today_col_idx) or 0,
                "months": [g(c) or 0 for c in month_cols_all]
            }
            rows.append(rec)
        r += 1
    return rows, month_headers

# ===== the builder you’ll call from run_stage1 =====
def build_by_agent_sheet_w90(
    out_path: str,
    private_pivot: str = "פיבוט פרטי",
    tedmiti_pivot: str = "פיבוט תדמיתי",
    sheet_name: str = "לפי סוכן",
    max_month_cols: int = 4,
) -> Tuple[bool, str, int]:

    wb_vals = load_workbook(out_path, data_only=True)
    if private_pivot not in wb_vals.sheetnames:
        wb_vals.close()
        return False, sheet_name, 0

    rows_all: Dict[Tuple[str, str], Dict] = {}  # (agent, channel) -> {F,G,months}
    month_headers: List[str] = []

    # --- פרטי
    r_p, h_p = _extract_from_pivot(wb_vals[private_pivot], max_month_cols=max_month_cols)
    for rec in r_p:
        rows_all[(rec["agent"], "שוק פרטי")] = {"F": rec["F"], "G": rec["G"], "months": rec["months"]}
    month_headers = list(h_p)

    # --- תדמיתי (אם קיים)
    if tedmiti_pivot in wb_vals.sheetnames:
        r_t, h_t = _extract_from_pivot(wb_vals[tedmiti_pivot], max_month_cols=max_month_cols)
        for rec in r_t:
            rows_all[(rec["agent"], "שוק תדמיתי")] = {"F": rec["F"], "G": rec["G"], "months": rec["months"]}
        for h in h_t:
            if h not in month_headers:
                month_headers.append(h)

    wb_vals.close()

    # סדר קבוצות קשיח כפי שביקשת
    private_groups = [
        ("שמעון כהן - מנהל אזור", ["אמנון ידידי","יואב מימון","משה כספי","עובדיה אבימלך"]),
        ("ישראל דנון- מנהל אזור", ["מרים בואזיזה","דוד פדלון","שמוליק מטרני","דניאל חורי"]),
        ("גיל רפאל", ["חאזם קדורה","חן בן דוד","לירון בן מוחה","גיא אלמוזנינו- סוכן","אוהד אסולין"]),
    ]
    private_singles = ["גילי סופר","בטי רובין"]  # אחריהם "סה\"כ מוקד", ואז "אריק יחזקאל" (פרטי), ואז "סיכום שוק פרטי"

    tedmiti_pair = ["יעל כץ","אריק יחזקאל"]  # אחריהם "סה\"כ" (של הזוג)
    ami_hachmon = ["משה רחמים","חיים שלו","ניר עזרא"]  # אחריהם "עמי חכמון"

    national = ["ארז ביתן","הילה אלסיאן- סחר","אלירן דהן","ליאור לוי - סחר","עינב כורם","מנהל אזור כללי"]  # אחריהם "סה\"כ רשתות ארציות"

    # --- יצירה/כתיבה
    wb = load_workbook(out_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = ["סוכן","ערוץ","פיגור גביה חודש נוכחי","פיגור גביה חודש קודם","הפרש", TOTAL_KEY, TODAY_KEY] + month_headers + ["סך פיגור"]
    ws.append(headers)

    # אינדקסים עמודות
    colA, colB = 1, 2
    colF, colG = 6, 7
    first_month_col = 8
    months_count = len(month_headers)
    col_sum_pigor = first_month_col + months_count  # העמודה שאחרי החודשים

    def _write_agent_row(name: str, channel: str) -> int:
        """כותב שורת סוכן (אם אין נתונים – כותב אפסים). מחזיר אינדקס שורה."""
        r = ws.max_row + 1
        ws.cell(row=r, column=colA).value = name
        ws.cell(row=r, column=colB).value = channel
        rec = rows_all.get((name, channel), {"F":0,"G":0,"months":[0]*months_count})
        # מספרים
        ws.cell(row=r, column=colF).value = float(rec["F"] or 0)
        ws.cell(row=r, column=colG).value = float(rec["G"] or 0)
        mv = list(rec["months"] or [])
        if len(mv) < months_count: mv += [0]*(months_count-len(mv))
        for j, val in enumerate(mv[:months_count]):
            ws.cell(row=r, column=first_month_col + j).value = float(val or 0)
        # סך פיגור + אחוזים
        if months_count:
            ws.cell(row=r, column=col_sum_pigor).value = f"=SUM({get_column_letter(first_month_col)}{r}:{get_column_letter(first_month_col+months_count-1)}{r})"
        else:
            ws.cell(row=r, column=col_sum_pigor).value = 0
        ws.cell(row=r, column=3).value = f"=IFERROR({get_column_letter(col_sum_pigor)}{r}/{get_column_letter(colF)}{r},0)"
        ws.cell(row=r, column=4).value = None
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        return r

    def _write_sum_row(title: str, source_rows: List[int]) -> int:
        """שורת סיכום: F,G וחודשים = סכימה ישירה של השורות; סך-פיגור = SUM של החודשים בשורת הסיכום; C= סך-פיגור/F; E=C-D"""
        r = ws.max_row + 1
        ws.cell(row=r, column=colA).value = title
        ws.cell(row=r, column=colB).value = ""
        # F, G
        # F (סה"כ סכום יתרת חוב)
        ws.cell(row=r, column=colF).value = ("=" + "+".join([f"{get_column_letter(colF)}{ri}" for ri in source_rows])) if source_rows else 0

        # G (סה"כ סכום יתרת חוב עד היום)
        ws.cell(row=r, column=colG).value = ("=" + "+".join([f"{get_column_letter(colG)}{ri}" for ri in source_rows])) if source_rows else 0

        # חודשי H..K (או כמה שיש)
        for j in range(months_count):
            col = first_month_col + j
            ws.cell(row=r, column=col).value = ("=" + "+".join([f"{get_column_letter(col)}{ri}" for ri in source_rows])) if source_rows else 0


        # סך פיגור (חיבור חודשי הסיכום), ואחוזים
        if months_count:
            ws.cell(row=r, column=col_sum_pigor).value = f"=SUM({get_column_letter(first_month_col)}{r}:{get_column_letter(first_month_col+months_count-1)}{r})"
        else:
            ws.cell(row=r, column=col_sum_pigor).value = 0
        ws.cell(row=r, column=3).value = f"=IFERROR({get_column_letter(col_sum_pigor)}{r}/{get_column_letter(colF)}{r},0)"
        ws.cell(row=r, column=4).value = None
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        # הדגשה לשורת סיכום (ויזואלי קל – בלי קווים עבים בשלב הלוגיקה)
        ws.cell(row=r, column=1).font = Font(bold=True)
        return r

    # ===== בנייה בפועל =====

    private_manager_sum_rows: List[int] = []
    # 1) שלוש קבוצות מנהלים (פרטי)
    for manager, members in private_groups:
        member_rows = []
        for nm in members:
            member_rows.append(_write_agent_row(nm, "שוק פרטי"))
        # שורת סיכום המנהל
        mgr_sum_r = _write_sum_row(manager, member_rows)
        private_manager_sum_rows.append(mgr_sum_r)

    # 2) יחידני פרטי: גילי, בטי -> "סה\"כ מוקד" -> אריק (פרטי)
    r_gili = _write_agent_row("גילי סופר", "שוק פרטי")
    r_bati = _write_agent_row("בטי רובין", "שוק פרטי")
    r_center = _write_sum_row("סה\"כ מוקד", [r_gili, r_bati])
    r_arik_p = _write_agent_row("אריק יחזקאל", "שוק פרטי")

    # 2.1) סיכום שוק פרטי = ישראל דנון + גיל רפאל + סה\"כ מוקד + אריק (פרטי)
    # (שמעון כהן לא נכנס לסיכום הזה לפי ההנחיות האחרונות שלך)
    # private_manager_sum_rows = [שמעון כהן, ישראל דנון, גיל רפאל] בסדר הבנייה
    # נבחר לפי שמות להימנע מבלבול:
    #   index 0 = שמעון, 1 = ישראל, 2 = גיל
    r_danon = private_manager_sum_rows[1]
    r_gil   = private_manager_sum_rows[2]
    r_private_total = _write_sum_row("סיכום שוק פרטי", [r_danon, r_gil, r_center, r_arik_p])

    # 3) זוג תדמיתי: יעל + אריק (תדמיתי) -> "סה\"כ"
    r_yael  = _write_agent_row("יעל כץ", "שוק תדמיתי")
    r_arik_t = _write_agent_row("אריק יחזקאל", "שוק תדמיתי")
    r_ted_pair_total = _write_sum_row("סה\"כ", [r_yael, r_arik_t])

    # 4) עמי חכמון: שלושה סוכנים (תדמיתי) -> "עמי חכמון"
    rows_ami = [_write_agent_row(nm, "שוק תדמיתי") for nm in ami_hachmon]
    r_ami_total = _write_sum_row("עמי חכמון", rows_ami)

    # 5) רשתות ארציות (בשלב הלוגיקה בלבד – נשים 0; אחר־כך נחבר נתונים מגיליונות המנהלים)
    rows_nat = []
    for nm in national:
        # ערוץ נשאר ריק או "רשתות ארציות" – לשיקולך. נשאיר ריק כדי לא לשבש פילטרים קיימים.
        r = _write_agent_row(nm, "")
        rows_nat.append(r)
    r_nat_total = _write_sum_row("סה\"כ רשתות ארציות", rows_nat)

    # 6) סה\"כ אחרון = סיכום שוק פרטי + סה\"כ (יעל+אריק תד') + עמי חכמון + סה\"כ רשתות
    _ = _write_sum_row("סה\"כ", [r_private_total, r_ted_pair_total, r_ami_total, r_nat_total])

    # פורמטים בסיסיים (C/D/E אחוזים)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = "0.00%"
        ws.cell(row=r, column=4).number_format = "0.00%"
        ws.cell(row=r, column=5).number_format = "0.00%"
        for c in range(6, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = "#,##0.00"

    # הקפאת שורה 1
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    # שמירה
    nrows = ws.max_row - 1
    wb.save(out_path); wb.close()
    return True, sheet_name, nrows



from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def link_national_from_manager_sheets_w90(out_path: str, sheet_name: str = "לפי סוכן") -> bool:
    """
    ממלא נתונים לרשתות ארציות ישירות מגיליונות המנהלים:
    F <- H, G <- I, וחודשי הפיגור לפי התאמת כותרות (לא לפי אינדקס!).
    גם קובע בעמודה B את 'רשתות ארציות'.
    """
    wb = load_workbook(out_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    # מיפוי כותרות ב'לפי סוכן'
    headers = {str(ws.cell(row=1, column=c).value or "").strip(): c
               for c in range(1, ws.max_column+1)}
    colA = headers.get("סוכן", 1)
    colB = headers.get("ערוץ", 2)
    colF = headers.get('סה"כ סכום יתרת חוב', 6)
    colG = headers.get('סה"כ סכום יתרת חוב עד היום', 7)

    # עמודות החודשים ב'לפי סוכן' לפי כותרת (עד לפני "סך פיגור")
    sum_pigor_col = None
    for name, idx in headers.items():
        if "סך פיגור" in str(name):
            sum_pigor_col = idx
    month_cols_in_by_agent = []
    month_titles_in_by_agent = []
    for name, idx in sorted(headers.items(), key=lambda kv: kv[1]):
        s = str(name or "")
        if any(h in s for h in ("חודש", "טרם")) and (sum_pigor_col is None or idx < sum_pigor_col):
            month_cols_in_by_agent.append(idx)
            month_titles_in_by_agent.append(s.strip())

    national_names = [
        "ארז ביתן",
        "הילה אלסיאן- סחר",
        "אלירן דהן",
        "ליאור לוי - סחר",
        "עינב כורם",
        "מנהל אזור כללי",
    ]

    # H,I אצל המנהל קבועים (לפי אישורך)
    mgr_F_col = 8  # H
    mgr_G_col = 9  # I

    def _set_lookup_last(r_target: int, c_target: int, tab: str, col_idx_in_tab: int):
        col_letter = get_column_letter(col_idx_in_tab)
        ws.cell(row=r_target, column=c_target).value = (
            f"=LOOKUP(2,1/('{tab}'!{col_letter}:{col_letter}<>\"\"),'{tab}'!{col_letter}:{col_letter})"
        )

    for nm in national_names:
        if nm not in wb.sheetnames:
            continue
        # מצא שורה בעמודה A
        row_idx = None
        for r in range(2, ws.max_row+1):
            if str(ws.cell(row=r, column=colA).value or "").strip() == nm:
                row_idx = r
                break
        if not row_idx:
            continue

        # קבע ערוץ = "רשתות ארציות"
        if colB:
            ws.cell(row=row_idx, column=colB).value = "רשתות ארציות"

        # F ו-G מתוך H ו-I של טאב המנהל (ערך אחרון לא ריק)
        _set_lookup_last(row_idx, colF, nm, mgr_F_col)
        _set_lookup_last(row_idx, colG, nm, mgr_G_col)

        # מיפוי חודשי פיגור לפי *כותרות*:
        ws_mgr = wb[nm]
        mgr_headers = {str(ws_mgr.cell(row=1, column=c).value or "").strip(): c
                       for c in range(1, ws_mgr.max_column+1)}

        for j, by_agent_month_col in enumerate(month_cols_in_by_agent):
            title = month_titles_in_by_agent[j]
            # מוצאים בטאב המנהל עמודה עם אותה כותרת בדיוק
            col_in_mgr = mgr_headers.get(title)
            if not col_in_mgr:
                # fallback: אם לא נמצא לפי שם – אל תקלקל, פשוט דלג
                continue
            _set_lookup_last(row_idx, by_agent_month_col, nm, col_in_mgr)

    wb.save(out_path); wb.close()
    return True



from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def rebind_all_sum_rows_w90(out_path: str, sheet_name: str = "לפי סוכן") -> bool:
    wb = load_workbook(out_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    # מיפוי כותרות
    hdr = {(ws.cell(row=1, column=c).value or "").strip(): c
           for c in range(1, ws.max_column+1)}
    colA = hdr.get("סוכן", 1)
    colB = hdr.get("ערוץ", 2)
    colF = hdr.get('סה"כ סכום יתרת חוב', 6)
    colG = hdr.get('סה"כ סכום יתרת חוב עד היום', 7)

    # עמודות חודשי פיגור (עד לפני 'סך פיגור')
    sum_pigor_col = hdr.get("סך פיגור", None)
    month_cols = []
    for name, idx in sorted(hdr.items(), key=lambda kv: kv[1]):
        s = str(name or "")
        if any(k in s for k in ("חודש", "טרם")) and (sum_pigor_col is None or idx < sum_pigor_col) and idx >= 8:
            month_cols.append(idx)
    months_count = len(month_cols)

    def find_row(label: str, channel: str|None=None) -> int|None:
        for r in range(2, ws.max_row+1):
            a = str(ws.cell(row=r, column=colA).value or "").strip()
            if a == label:
                if channel is None:
                    return r
                b = str(ws.cell(row=r, column=colB).value or "").strip()
                if b == channel:
                    return r
        return None

    def write_sum_to_row(r_target: int, src_rows: list[int]):
        if not r_target: return
        # F,G סכומי מקור
        ws.cell(row=r_target, column=colF).value = ("=" + "+".join([f"{get_column_letter(colF)}{ri}" for ri in src_rows])) if src_rows else 0
        ws.cell(row=r_target, column=colG).value = ("=" + "+".join([f"{get_column_letter(colG)}{ri}" for ri in src_rows])) if src_rows else 0
        # חודשי פיגור (H..)
        for c in month_cols:
            ws.cell(row=r_target, column=c).value = ("=" + "+".join([f"{get_column_letter(c)}{ri}" for ri in src_rows])) if src_rows else 0

    # קבוצות/חברים לפי ההגדרה הקשיחה שבבנאי
    private_groups = {
        "שמעון כהן - מנהל אזור": ["אמנון ידידי","יואב מימון","משה כספי","עובדיה אבימלך"],
        "ישראל דנון- מנהל אזור": ["מרים בואזיזה","דוד פדלון","שמוליק מטרני","דניאל חורי"],
        "גיל רפאל": ["חאזם קדורה","חן בן דוד","לירון בן מוחה","גיא אלמוזנינו- סוכן","אוהד אסולין"],
    }
    # 1) מנהלים (פרטי)
    for mgr, members in private_groups.items():
        r_mgr = find_row(mgr, None)
        src = [find_row(m, "שוק פרטי") for m in members]
        src = [r for r in src if r]
        write_sum_to_row(r_mgr, src)

    # 2) סה"כ מוקד = גילי+בטי (פרטי)
    r_center = find_row('סה"כ מוקד')
    r_gili = find_row("גילי סופר", "שוק פרטי")
    r_bati = find_row("בטי רובין", "שוק פרטי")
    write_sum_to_row(r_center, [r for r in (r_gili, r_bati) if r])

    # 3) סיכום שוק פרטי = ישראל דנון + גיל רפאל + סה"כ מוקד + אריק (פרטי)
    r_private_total = find_row("סיכום שוק פרטי")
    r_danon = find_row("ישראל דנון- מנהל אזור")
    r_gil   = find_row("גיל רפאל")
    r_arik_p = find_row("אריק יחזקאל", "שוק פרטי")
    write_sum_to_row(r_private_total, [r for r in (r_danon, r_gil, r_center, r_arik_p) if r])

    # 4) סה"כ (זוג תדמיתי) = יעל + אריק (תדמיתי)
    # r_ted_total = None
    # # נאתר "סה\"כ" של הזוג לפי ערוץ "שוק תדמיתי"
    # for r in range(2, ws.max_row+1):
    #     if str(ws.cell(row=r, column=colA).value or "").strip() == 'סה"כ' and str(ws.cell(row=r, column=colB).value or "").strip() == "שוק תדמיתי":
    #         r_ted_total = r; break
    # r_yael = find_row("יעל כץ", "שוק תדמיתי")
    # r_arik_t = find_row("אריק יחזקאל", "שוק תדמיתי")
    # write_sum_to_row(r_ted_total, [r for r in (r_yael, r_arik_t) if r])


#### סה"כ שוק תדמיתי יחדים, הבלוק למעלה לא היה מחשב שורות נכונות
    r_ted_total = None
    # נסיון 1: A='סה"כ' ו-B='שוק תדמיתי' (אם הכתרת את הערוץ)
    for r in range(2, ws.max_row+1):
        a = str(ws.cell(row=r, column=colA).value or "").strip()
        b = str(ws.cell(row=r, column=colB).value or "").strip()
        if a == 'סה"כ' and b == 'שוק תדמיתי':
            r_ted_total = r
        break
    
    if r_ted_total is None:
        for r in range(2, ws.max_row+1):
            a = str(ws.cell(row=r, column=colA).value or "").strip()
            b = str(ws.cell(row=r, column=colB).value or "").strip()
            if a == 'סה"כ' and b == '':
                prev_names = {
                    str(ws.cell(row=r-1, column=colA).value or "").strip(),
                    str(ws.cell(row=r-2, column=colA).value or "").strip()
                }
                prev_channels = {
                    str(ws.cell(row=r-1, column=colB).value or "").strip(),
                    str(ws.cell(row=r-2, column=colB).value or "").strip()
                }
                if {'יעל כץ', 'אריק יחזקאל'}.issubset(prev_names) and 'שוק תדמיתי' in prev_channels:
                    r_ted_total = r
                    break
                
    r_yael   = find_row("יעל כץ", "שוק תדמיתי")
    r_arik_t = find_row("אריק יחזקאל", "שוק תדמיתי")
    write_sum_to_row(r_ted_total, [r for r in (r_yael, r_arik_t) if r])
    
    #הסוף של בלוק החדש של סה"כ שוק תדמיתי יחידים

    # 5) עמי חכמון (תדמיתי)
    r_ami = find_row("עמי חכמון")
    ami_members = ["משה רחמים","חיים שלו","ניר עזרא"]
    src_ami = [find_row(m, "שוק תדמיתי") for m in ami_members]
    src_ami = [r for r in src_ami if r]
    write_sum_to_row(r_ami, src_ami)

    # 6) סה"כ רשתות ארציות
    r_nat_total = find_row('סה"כ רשתות ארציות')
    nat_members = ["ארז ביתן","הילה אלסיאן- סחר","אלירן דהן","ליאור לוי - סחר","עינב כורם","מנהל אזור כללי"]
    src_nat = [find_row(m, None) for m in nat_members]
    src_nat = [r for r in src_nat if r]
    write_sum_to_row(r_nat_total, src_nat)

    # 7) סה"כ אחרון = סיכום שוק פרטי + סה"כ (תדמיתי) + עמי חכמון + סה"כ רשתות
    r_master = None
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(row=r, column=colA).value or "").strip() == 'סה"כ':
            r_master = r; break
    src_master = [r for r in (r_private_total, r_ted_total, r_ami, r_nat_total) if r]
    write_sum_to_row(r_master, src_master)

    wb.save(out_path); wb.close()
    return True



from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def ensure_pigor_sum_and_pct_w90(xlsx_path: str, sheet_name: str = "לפי סוכן") -> bool:
    """
    מעדכן *רק*:
      L = SUM(עמודות חודשי הפיגור)
      C = IFERROR(L/F,0)
      E = C - D
    + פורמטים (C/D/E אחוזים; מ-F והלאה מספרים).
    לא נוגע בכלום אחר.
    """
    wb = load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    # מיפוי כותרות
    hdr = {(ws.cell(row=1, column=c).value or "").strip(): c
           for c in range(1, ws.max_column + 1)}

    col_total = hdr.get('סה"כ סכום יתרת חוב')              # F
    col_today = hdr.get('סה"כ סכום יתרת חוב עד היום')     # G
    col_pigor = hdr.get('סך פיגור')                        # L (אם חסרה – ניצור)

    # 1) איתור עמודות החודשים
    # --- במקום הבלוק שמוצא month_cols, שים את זה ---
    BAD_TITLES = {
    "שיטת תשלום לקוח משלם", "שיטת תשלום", "לקוח משלם",
    "פיגור גביה חודש נוכחי", "פיגור גביה חודש קודם", "הפרש",
    }

# 1) קודם לפי שם (מוציאים רעים), ואז נשמור רק טורים מימין ל-G (>=8)
    month_cols = sorted({
        c for name, c in hdr.items()
        if any(k in str(name) for k in ("חודש", "טרם"))
        and str(name).strip() not in BAD_TITLES
        and c >= 8
    })

# 2) אם לא נמצאו לפי שם — fallback לפי מיקום (אחרי G ועד לפני 'סך פיגור')
    if not month_cols:
        start = max((col_today or 7) + 1, 8)  # H=8
        end   = (col_pigor - 1) if col_pigor else ws.max_column
        if end >= start:
            month_cols = list(range(start, end + 1))

    # 3) ביטוח נוסף: תוריד גם מה-fallback כותרות בעייתיות אם במקרה קיימות
    month_cols = [
        c for c in month_cols
        if str(ws.cell(row=1, column=c).value or "").strip() not in BAD_TITLES
    ]
    if not month_cols:
        wb.save(xlsx_path); wb.close(); return False


    first_m, last_m = month_cols[0], month_cols[-1]

    # 2) אם אין "סך פיגור" – ניצור עמודה אחרי החודש האחרון
    if not col_pigor:
        col_pigor = last_m + 1
        ws.insert_cols(col_pigor)
        ws.cell(row=1, column=col_pigor).value = "סך פיגור"

    # 3) כתיבת נוסחאות לכל השורות (2..max_row)
    Lcol_letter = get_column_letter(col_pigor)
    firstM_letter = get_column_letter(first_m)
    lastM_letter  = get_column_letter(last_m)
    F_letter = get_column_letter(col_total) if col_total else None

    for r in range(2, ws.max_row + 1):
        # דלג על שורה ריקה לגמרי (בטיחות)
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ws.max_column + 1)):
            continue

        # L = SUM(H..K) (לפי הטווח שמצאנו)
        ws.cell(row=r, column=col_pigor).value = f"=SUM({firstM_letter}{r}:{lastM_letter}{r})"
        ws.cell(row=r, column=col_pigor).number_format = "#,##0.00"

        # C = IFERROR(L/F,0) — רק אם יש F
        if F_letter:
            ws.cell(row=r, column=3).value = f"=IFERROR({Lcol_letter}{r}/{F_letter}{r},0)"
        ws.cell(row=r, column=3).number_format = "0.00%"

        # D בפורמט אחוז (ערך יישאר כפי שהוא)
        ws.cell(row=r, column=4).number_format = "0.00%"

        # E = C - D
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        ws.cell(row=r, column=5).number_format = "0.00%"

        # פורמט מספרי מ-F והלאה
        for c in range(6, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = "#,##0.00"

    wb.save(xlsx_path); wb.close()
    return True





from openpyxl.styles import PatternFill, Font, Border, Side

def ensure_group_blank_rows_w90(xlsx_path: str, sheet_name: str = "לפי סוכן") -> bool:
    targets = {"גילי סופר", "יעל כץ", "ארז ביתן", "סיכום שוק פרטי"}
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    r = 2
    while r <= ws.max_row:
        name = str(ws.cell(row=r, column=1).value or "").strip()
        if name in targets:
            prev_is_empty = False
            if r > 2:
                prev_is_empty = all((ws.cell(row=r-1, column=c).value in (None, "")) for c in range(1, ws.max_column+1))
            if not prev_is_empty:
                ws.insert_rows(r, 1)
                # לא מדלגים פעמיים—אחרי ההוספה נתקדם עם r+=2 כדי לא להיכנס ללופ אין-סופי
                r += 1
        r += 1

    wb.save(xlsx_path); wb.close()
    return True



def remove_thick_above_national_total(xlsx_path: str, sheet_name: str = "לפי סוכן") -> bool:
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    # חפש את השורה של "סה\"כ רשתות ארציות" בעמודה A
    target_row = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if str(a or "").strip() == 'סה"כ רשתות ארציות':
            target_row = r
            break
    if not target_row or target_row <= 2:
        wb.close(); return False

    # ננקה את הקו העבה מהשורה שמעל (r-1) בכל הטורים
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=target_row - 1, column=c)
        b = cell.border
        # הסרת ה-bottom העבה (משאירים את השאר כמו שהם)
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=Side(style=None))
    wb.save(xlsx_path); wb.close()
    return True



def style_groups_colA_only_w90(xlsx_path: str, sheet_name: str = "לפי סוכן") -> bool:
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    fills = [
        PatternFill("solid", start_color="FFE6F4EA", end_color="FFE6F4EA"),  # ירוק בהיר
        PatternFill("solid", start_color="FFFFF2CC", end_color="FFFFF2CC"),  # צהוב בהיר
        PatternFill("solid", start_color="FFE8F4FD", end_color="FFE8F4FD"),  # תכלת
    ]
    thick = Side(style="thick", color="000000")

    # סוף-קבוצה: מוסיפים מפורשות את "גיל רפאל"
    def is_group_end(label: str) -> bool:
        s = (label or "").strip()
        return (
            s in {"גיל רפאל", "עמי חכמון", "סה\"כ רשתות ארציות", "סה\"כ", "סיכום שוק פרטי", "סה\"כ מוקד"} or
            ("מנהל אזור" in s) or ("מנהלת אזור" in s)
        )

    # Bold רק לשורות שביקשת
    BOLD_TITLES = {
        "שמעון כהן - מנהל אזור",
        "ישראל דנון- מנהל אזור",
        "גיל רפאל",
        "עמי חכמון",
        "סה\"כ רשתות ארציות",
        "סה\"כ",
    }

    group_start = 2
    group_index = 0  # נספור קבוצות שנצבעו
    r = 2
    while r <= ws.max_row:
        name = ws.cell(row=r, column=1).value
        if is_group_end(name):
            # צבע עמודה A לכל השורות בקבוצה – אבל רק ל-3 הקבוצות הראשונות
            if group_index < 3:
                fill = fills[group_index]  # 0..2
                for rr in range(group_start, r+1):
                    ws.cell(row=rr, column=1).fill = fill

            # קו עבה מתחת לשורת הסיכום + Bold לפי הרשימה
            for c in range(1, ws.max_column+1):
                cell = ws.cell(row=r, column=c)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=thick
                )
                if str(name or "").strip() in BOLD_TITLES:
                    cell.font = Font(bold=True)

            group_index += 1
            group_start = r + 1
        r += 1

    # פורמטים (ליתר ביטחון): C/D/E אחוזים, שאר העמודות כספרות
    for rr in range(2, ws.max_row + 1):
        for cc in (3,4,5):
            ws.cell(row=rr, column=cc).number_format = "0.00%"
        for cc in range(6, ws.max_column + 1):
            ws.cell(row=rr, column=cc).number_format = "#,##0.00"

    wb.save(xlsx_path); wb.close()
    return True




from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def set_column_layout_w90(xlsx_path: str, sheet_name: str = "לפי סוכן") -> bool:
    wb = load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet_name]

    # מיפוי כותרות לשימוש נוח
    hdr = {(ws.cell(row=1, column=c).value or "").strip(): c
           for c in range(1, ws.max_column + 1)}

    def set_w(col_idx: int, width: float):
        if not col_idx:
            return
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # רוחבים קבועים לעמודות מפתח
    set_w(hdr.get("סוכן"), 24)
    set_w(hdr.get("ערוץ"), 12)

    # C,D,E (אחוזים)
    set_w(3, 12); set_w(4, 12); set_w(5, 12)

    # F,G סכומים עיקריים
    set_w(hdr.get('סה"כ סכום יתרת חוב'), 18)
    set_w(hdr.get('סה"כ סכום יתרת חוב עד היום'), 18)

    # עמודות חודשים (כולל "טרם...")
    for name, c in hdr.items():
        s = str(name)
        if ("חודש" in s) or ("טרם" in s):
            set_w(c, 16)

    # "סך פיגור"
    set_w(hdr.get("סך פיגור"), 18)

    # התאמת יישור ו-shrink-to-fit לכל המספרים (C..סוף), אחוזים במרכז/מספרים מימין
    last_col = ws.max_column
    for r in range(2, ws.max_row + 1):
        for c in range(3, last_col + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal=("center" if c in (3, 4, 5) else "right"),
                vertical="center",
                shrink_to_fit=True,
                wrap_text=False
            )

    # כותרות ו-RTL
    for c in range(1, last_col + 1):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    wb.save(xlsx_path); wb.close()
    return True
