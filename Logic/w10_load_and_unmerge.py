import os
import openpyxl

def load_and_unmerge(src_path: str, sheet_hint: str = "sheet1", temp_out_dir: str = None) -> str:
    if temp_out_dir is None:
        temp_out_dir = os.path.dirname(src_path) or "."
    os.makedirs(temp_out_dir, exist_ok=True)
    tmp_path = os.path.join(temp_out_dir, "_temp_unmerged.xlsx")

    wb = openpyxl.load_workbook(src_path, data_only=True)
    target_name = None
    for name in wb.sheetnames:
        if name.lower() == sheet_hint.lower():
            target_name = name
            break
    if target_name is None:
        target_name = wb.sheetnames[0]
    ws = wb[target_name]

    merged_ranges = list(ws.merged_cells.ranges)
    for mrange in merged_ranges:
        min_row, min_col, max_row, max_col = mrange.min_row, mrange.min_col, mrange.max_row, mrange.max_col
        val = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = val

    wb.save(tmp_path)
    return tmp_path
