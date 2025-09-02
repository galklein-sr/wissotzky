import re
from typing import List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from Logic.w74_helper_suppression import suppress_rows_by_helper


SUM_ANCHOR_AFTER_TODAY = 'סה"כ סכום יתרת חוב עד היום'   # I
SUM_ANCHOR_TOTAL      = 'סה"כ סכום יתרת חוב'            # H
HELPER_COL_NAME       = "טור עזר"                        # N
FORBIDDEN = r'[:\\/?*\[\]]'

def _pick_first(df_cols: List[str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df_cols:
            return c
    return None

def _dynamic_month_cols(df_cols: List[str], anchor: str, max_cols: int = 4) -> List[str]:
    if anchor in df_cols:
        idx = df_cols.index(anchor)
        return [c for c in df_cols[idx+1 : idx+1+max_cols] if c in df_cols]
    return []

def _autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None: continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def _style_header(ws):
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for j in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    try: ws.sheet_view.rightToLeft = True
    except Exception: pass

def _add_column_sums_row(ws, header_names: List[str], total_col: str, today_col: str):
    """
    מוסיף שורת סכומים לכל העמודות מ-H עד N (כולל) שתי שורות מתחת לשורה האחרונה.
    מניח סדר: ... G('קוד סוכן'), H(total_col), I(today_col), J..M(dyn), N(HELPER_COL_NAME).
    """
    if total_col not in header_names or today_col not in header_names or HELPER_COL_NAME not in header_names:
        return

    name_to_idx0 = {name: i for i, name in enumerate(header_names)}
    first_sum_idx0 = name_to_idx0[total_col]       # H (0-based)
    last_sum_idx0  = name_to_idx0[HELPER_COL_NAME] # N (0-based)
    if last_sum_idx0 < first_sum_idx0:
        return

    last_row = ws.max_row
    sum_row  = last_row + 2  # שתי שורות מתחת לשורה האחרונה

    # "סכום :" בעמודה שלפני H -> כלומר G (0-based H => label col = H-1 => first_sum_idx0)
    label_col_excel = max(1, first_sum_idx0)
    ws.cell(row=sum_row, column=label_col_excel).value = "סכום :"
    ws.cell(row=sum_row, column=label_col_excel).font = Font(bold=True)

    start_data_row = 2  # אחרי שורת כותרת
    for idx0 in range(first_sum_idx0, last_sum_idx0 + 1):
        col_excel = idx0 + 1
        col_letter = get_column_letter(col_excel)
        ws.cell(row=sum_row, column=col_excel).value = f"=SUM({col_letter}{start_data_row}:{col_letter}{last_row})"
        ws.cell(row=sum_row, column=col_excel).number_format = '#,##0.00'
        ws.cell(row=sum_row, column=col_excel).font = Font(bold=True)

def _drop_total_like_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    מסיר:
    1) שורת 'סכום' טקסטואלית (אם נותרה בטעות).
    2) שורות סכום-מספריות (כל עמודות המזהים ריקות, אבל ב-H יש מספר).
    """
    cols = list(df.columns)
    total_col = SUM_ANCHOR_TOTAL if SUM_ANCHOR_TOTAL in cols else None
    today_col = SUM_ANCHOR_AFTER_TODAY if SUM_ANCHOR_AFTER_TODAY in cols else None

    out = df.copy()

    # 1) להסיר טקסט "סכום" אם איכשהו נשאר
    if total_col:
        out = out[out[total_col].astype(str).str.strip() != "סכום"]
    if today_col:
        out = out[out[today_col].astype(str).str.strip() != "סכום"]

    # 2) להסיר שורה "מספרית" ללא מזהים (שורת סכום של H מהמעובד)
    id_cols = [c for c in ["מנהל סחר","מנהל אזור","מנהל איזור","סוכן","ערוץ",
                           "קוד סוכן","קוד לקוח קצה","קוד לקוח משלם","לקוח קצה","לקוח משלם"] if c in cols]
    if total_col:
        def _is_empty(x): 
            return (x.isna()) | (x.astype(str).str.strip() == "")
        mask_ids_empty = out[id_cols].apply(lambda s: _is_empty(s), axis=0).all(axis=1) if id_cols else False
        mask_total_num = pd.to_numeric(out[total_col], errors="coerce").notna()
        # נסיר רק שורות שהן גם ללא מזהים וגם יש בהן מספר ב-H
        out = out[~(mask_ids_empty & mask_total_num)]

    return out

# ---------- main builder ----------
def build_region_general_full_columns(
    processed_df: pd.DataFrame,
    output_path: str,
    sheet_name: str = "מנהל אזור כללי",
    max_month_cols_after_today: int = 4
) -> Tuple[bool, str]:
    """
    בונה גיליון 'מנהל אזור כללי' בפורמט זהה לגיליונות מנהלי סחר:
    A: מנהל סחר
    B: מנהל אזור
    C: סוכן
    D: ערוץ
    E: קוד לקוח משלם   (ממפה מקצה/משלם)
    F: לקוח משלם       (ממפה מקצה/משלם)
    G: קוד סוכן
    H: סה"כ סכום יתרת חוב
    I: סה"כ סכום יתרת חוב עד היום
    J..M: עד 4 חודשים דינמיים שאחרי I
    N: טור עזר = סכום שורה של J..M
    + שורת סכום לכל H..N
    """
    df = _drop_total_like_rows(processed_df.copy())
    cols = list(df.columns)

    region_col      = _pick_first(cols, ["מנהל אזור", "מנהל איזור"])
    payer_code_src  = _pick_first(cols, ["קוד לקוח קצה", "קוד לקוח משלם"])
    payer_name_src  = _pick_first(cols, ["לקוח קצה", "לקוח משלם"])
    agent_code_col  = "קוד סוכן" if "קוד סוכן" in cols else None
    agent_col       = "סוכן" if "סוכן" in cols else None
    channel_col     = "ערוץ" if "ערוץ" in cols else None

    total_col = SUM_ANCHOR_TOTAL if SUM_ANCHOR_TOTAL in cols else SUM_ANCHOR_TOTAL
    today_col = SUM_ANCHOR_AFTER_TODAY if SUM_ANCHOR_AFTER_TODAY in cols else SUM_ANCHOR_AFTER_TODAY

    dyn_cols = _dynamic_month_cols(cols, SUM_ANCHOR_AFTER_TODAY, max_month_cols_after_today)
    if not dyn_cols and SUM_ANCHOR_TOTAL in cols:
        dyn_cols = _dynamic_month_cols(cols, SUM_ANCHOR_TOTAL, max_month_cols_after_today)

    # סדר ונתונים
    order, data = [], {}

    # A..D
    order += ["מנהל סחר", "מנהל אזור", "סוכן", "ערוץ"]
    data["מנהל סחר"] = df["מנהל סחר"] if "מנהל סחר" in cols else pd.Series([], dtype=object)
    data["מנהל אזור"] = df[region_col] if region_col else pd.Series([], dtype=object)
    data["סוכן"] = df[agent_col] if agent_col else pd.Series([], dtype=object)
    data["ערוץ"] = df[channel_col] if channel_col else pd.Series([], dtype=object)

    # E..F (משלם במקום קצה כשאפשר)
    order += ["קוד לקוח משלם", "לקוח משלם"]
    data["קוד לקוח משלם"] = df[payer_code_src] if payer_code_src else pd.Series([], dtype=object)
    data["לקוח משלם"] = df[payer_name_src] if payer_name_src else pd.Series([], dtype=object)

    # G
    order += ["קוד סוכן"]
    data["קוד סוכן"] = df[agent_code_col] if agent_code_col else pd.Series([], dtype=object)

    # H..I
    order += [total_col, today_col]
    data[total_col] = pd.to_numeric(df[SUM_ANCHOR_TOTAL], errors="coerce") if SUM_ANCHOR_TOTAL in cols else pd.Series([], dtype=float)
    data[today_col] = pd.to_numeric(df[SUM_ANCHOR_AFTER_TODAY], errors="coerce") if SUM_ANCHOR_AFTER_TODAY in cols else pd.Series([], dtype=float)

    # J..M
    for c in dyn_cols:
        order.append(c)
        data[c] = pd.to_numeric(df[c], errors="coerce") if c in cols else pd.Series([], dtype=float)

    # N: טור עזר
    order.append(HELPER_COL_NAME)
    if dyn_cols and len(df):
        dyn_num = pd.concat([pd.to_numeric(df[c], errors="coerce") for c in dyn_cols], axis=1)
        data[HELPER_COL_NAME] = dyn_num.sum(axis=1, skipna=True)
    else:
        data[HELPER_COL_NAME] = pd.Series([0.0]*len(df), dtype=float)

    out_df = pd.DataFrame(data)[order]
    
    # דיכוי שורות לפי טור עזר < -1000
    cols_out = out_df.columns.tolist()
    today = SUM_ANCHOR_AFTER_TODAY if SUM_ANCHOR_AFTER_TODAY in cols_out else SUM_ANCHOR_AFTER_TODAY
    dyn_for_suppress = _dynamic_month_cols(cols_out, today, max_month_cols_after_today)
    out_df = suppress_rows_by_helper(out_df, month_cols=dyn_for_suppress, helper_col=HELPER_COL_NAME, threshold=-1000)


    # כתיבה לקובץ
    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    ws.append(out_df.columns.tolist())
    for _, row in out_df.iterrows():
        ws.append(row.tolist())

    _style_header(ws)
    _add_column_sums_row(
        ws,
        header_names=out_df.columns.tolist(),
        total_col=total_col,
        today_col=today_col
    )
    _autosize(ws)
    

    wb.save(output_path)
    return True, sheet_name