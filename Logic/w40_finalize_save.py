import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from .utils import ts_now, base_from_path, safe_mkdir
from .headers_stage1 import AMOUNT_HEADERS

def _autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def _apply_header_style(ws):
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

def _apply_number_formats(ws):
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for name in AMOUNT_HEADERS:
        col = header_map.get(name)
        if not col:
            continue
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

def save_processed(df: pd.DataFrame, input_path: str, output_dir: str, sheet_name: str = "מעובד") -> str:
    safe_mkdir(output_dir)
    out_name = f"{base_from_path(input_path)}_{ts_now()}.xlsx"
    out_path = os.path.join(output_dir, out_name)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        _apply_header_style(ws)
        _autosize(ws)
        _apply_number_formats(ws)

    return out_path