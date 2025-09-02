from openpyxl import load_workbook
from typing import Tuple, List

REGION_HEADERS = ("מנהל אזור", "מנהל איזור")

def _find_region_col(headers: List[str]) -> int:
    """החזר אינדקס 1-based של עמודת מנהל אזור/איזור, או 0 אם לא נמצאה."""
    for name in REGION_HEADERS:
        try:
            return headers.index(name) + 1
        except ValueError:
            continue
    return 0

def refine_private_region_rows(
    xlsx_path: str,
    sheet_name: str = "שוק פרטי",
    forbidden_substr: str = "רפי מור יוסף"
) -> Tuple[bool, int]:
    """
    מוחק מגיליון 'שוק פרטי' כל שורה שבה בעמודת מנהל אזור/איזור מופיע 'רפי מור יוסף'.
    מחזיר (בוצע_שינוי, כמות_מחיקות).
    """
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        return False, 0

    ws = wb[sheet_name]
    if ws.max_row < 2 or ws.max_column < 1:
        return False, 0

    headers = [ (ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1) ]
    region_col = _find_region_col(headers)
    if region_col == 0:
        return False, 0

    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=region_col).value
        s = "" if val is None else str(val).strip()
        if forbidden_substr in s:
            rows_to_delete.append(r)

    if not rows_to_delete:
        return False, 0

    # מוחקים מלמטה למעלה לשמירת אינדקסים
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    wb.save(xlsx_path)
    return True, len(rows_to_delete)
