from typing import List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import re

SUM_ANCHOR_AFTER_TODAY = 'סה"כ סכום יתרת חוב עד היום'   # I
SUM_ANCHOR_TOTAL      = 'סה"כ סכום יתרת חוב'            # H


def _norm_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace('\u200f','').replace('\u200e','')
    s = s.replace('–','-').replace('—','-').replace('-','-').replace('־','-')
    s = re.sub(r'\s*-\s*', ' - ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

# ---------- helpers ----------
def _pick_first(df_cols: List[str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df_cols:
            return c
    return None

def _dynamic_month_cols(df_cols: List[str], anchor: str, max_cols: int = 4) -> List[str]:
    if anchor in df_cols:
        idx = df_cols.index(anchor)
        return [c for c in df_cols[idx+1 : idx+1+max_cols] if c in df_cols]
    return []

def _style_header(ws):
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for j in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

def _autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def _add_total_row(ws, first_numeric_col_idx: int = 2):
    """
    מוסיף שורת סכום לכל העמודות המספריות (מעמודה 2 = B ועד סוף)
    ושם את הטקסט 'סכום :' בעמודה הראשונה (A) באותה שורה.
    """
    last_row = ws.max_row
    if last_row < 2:
        return
    sum_row = last_row + 2  # שתי שורות אחרי הנתונים
    ws.cell(row=sum_row, column=1).value = "סכום :"
    ws.cell(row=sum_row, column=1).font = Font(bold=True)

    start_data = 2  # אחרי הכותרת
    for col in range(first_numeric_col_idx, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.cell(row=sum_row, column=col).value = f"=SUM({col_letter}{start_data}:{col_letter}{last_row})"
        ws.cell(row=sum_row, column=col).number_format = '#,##0.00'
        ws.cell(row=sum_row, column=col).font = Font(bold=True)

def _build_pivot_by_agent(df: pd.DataFrame, max_month_cols_after_today: int = 4) -> Optional[pd.DataFrame]:
    """
    בונה DF מסוכם ברמת 'סוכן' לעמודות: H, I, ו-J..M (דינמי).
    כולל סוכן ריק ("(ריק)") כדי להימנע מ-None.
    """
    if df.empty:
        return None

    cols = list(df.columns)
    agent_col = "סוכן" if "סוכן" in cols else None
    if not agent_col:
        return None

    total_col = SUM_ANCHOR_TOTAL if SUM_ANCHOR_TOTAL in cols else None
    today_col = SUM_ANCHOR_AFTER_TODAY if SUM_ANCHOR_AFTER_TODAY in cols else None
    if not total_col and not today_col:
        return None

    dyn_cols = _dynamic_month_cols(cols, SUM_ANCHOR_AFTER_TODAY, max_month_cols_after_today)
    if not dyn_cols and total_col:
        dyn_cols = _dynamic_month_cols(cols, SUM_ANCHOR_TOTAL, max_month_cols_after_today)

    sum_cols: List[str] = []
    if total_col: sum_cols.append(total_col)
    if today_col: sum_cols.append(today_col)
    sum_cols += dyn_cols

    # המרה למספרים
    for c in sum_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # אל תזרוק סוכן ריק — תן לו תווית "(ריק)"
    sub = df.copy()
    sub[agent_col] = sub[agent_col].astype(str).str.strip()
    sub.loc[sub[agent_col] == "", agent_col] = "(ריק)"

    pivot = sub.groupby(agent_col, dropna=False)[sum_cols].sum(min_count=1).reset_index()

    ordered = [agent_col]
    if total_col: ordered.append(total_col)
    if today_col: ordered.append(today_col)
    ordered += [c for c in dyn_cols if c in sum_cols]
    pivot = pivot[ordered]

    return pivot


def _write_pivot_to_sheet(out_path: str, sheet_name: str, pivot_df: pd.DataFrame):
    wb = load_workbook(out_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # כתיבת כותרות
    ws.append(pivot_df.columns.tolist())
    # כתיבת שורות
    for _, row in pivot_df.iterrows():
        ws.append(row.tolist())

    _style_header(ws)
    _add_total_row(ws, first_numeric_col_idx=2)
    _autosize(ws)
    wb.save(out_path)

# ---------- builders ----------
def build_pivot_private(
    df_processed: pd.DataFrame,
    output_path: str,
    manager_name: str = "רפי מור יוסף-סחר",
    channel_value: str = "שוק פרטי",
    sheet_name: str = "פיבוט פרטי",
    max_month_cols_after_today: int = 4
) -> Tuple[bool, str]:
    """
    פיבוט פרטי: מסנן לפי מנהל סחר + ערוץ, מסכם לפי 'סוכן'.
    """
    if df_processed.empty:
        return False, sheet_name

    cols = list(df_processed.columns)
    mgr_col = "מנהל סחר" if "מנהל סחר" in cols else None
    ch_col  = "ערוץ" if "ערוץ" in cols else None
    if not mgr_col or not ch_col:
        return False, sheet_name
    
    
    mgr_series_norm = df_processed[mgr_col].astype(str).map(_norm_text)
    ch_series_norm  = df_processed[ch_col].astype(str).map(_norm_text)
    target_mgr = _norm_text(manager_name)
    target_ch  = _norm_text(channel_value)

    mask = (mgr_series_norm == target_mgr) & (ch_series_norm == target_ch)
    sub = df_processed.loc[mask].copy()
    
    

    pivot_df = _build_pivot_by_agent(sub, max_month_cols_after_today=max_month_cols_after_today)
    if pivot_df is None or pivot_df.empty:
        return False, sheet_name

    _write_pivot_to_sheet(output_path, sheet_name, pivot_df)
    return True, sheet_name

def build_pivot_tedmiti(
    df_processed: pd.DataFrame,
    output_path: str,
    manager_name: str = "עמי חכמון",
    sheet_name: str = "פיבוט תדמיתי",
    max_month_cols_after_today: int = 4
) -> Tuple[bool, str]:
    """
    פיבוט תדמיתי: מסנן לפי מנהל סחר (ללא סינון ערוץ), מסכם לפי 'סוכן'.
    """
    if df_processed.empty:
        return False, sheet_name

    cols = list(df_processed.columns)
    mgr_col = "מנהל סחר" if "מנהל סחר" in cols else None
    if not mgr_col:
        return False, sheet_name

    sub = df_processed.loc[
        (df_processed[mgr_col].astype(str).str.strip() == manager_name)
    ].copy()

    pivot_df = _build_pivot_by_agent(sub, max_month_cols_after_today=max_month_cols_after_today)
    if pivot_df is None or pivot_df.empty:
        return False, sheet_name

    _write_pivot_to_sheet(output_path, sheet_name, pivot_df)
    return True, sheet_name
