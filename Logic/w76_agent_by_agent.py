import re
from typing import Dict, List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# שמות אפשריים לעמודת סוכן בפיבוט
AGENT_KEYS = ["תוויות שורה", "סוכן", "Row Labels"]

# ביטויים לחיפוש עמודות סכומים
TOTAL_KEY = "סה\"כ סכום יתרת חוב"
TODAY_KEY = "סה\"כ סכום יתרת חוב עד היום"

# זיהוי עמודות חודשים (דינמי)
MONTH_HINTS = ["חודש", "טרם"]

SHEET_PIVOT_PRIVATE_DEFAULT = "פיבוט פרטי"
SHEET_BY_AGENT_DEFAULT      = "לפי סוכן"

PERCENT_FMT = "0.00%"
NUMBER_FMT  = "#,##0.00"

def _style_header(ws: Worksheet):
    fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for j in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

def _autosize(ws: Worksheet):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

def _row_values(ws: Worksheet, r: int) -> List[str]:
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        out.append("" if v is None else str(v).strip())
    return out

def _find_header_row(ws: Worksheet, scan_rows: int = 30) -> Optional[int]:
    # קודם מנסה למצוא לפי עמודה A אם יש אחד מהמפתחות לסוכן
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        if str(a).strip() in AGENT_KEYS:
            return r
    # אם לא נמצא, נחפש שורה "עשירה" שכבר כוללת כותרות פיבוט (agent/total/today/חודש)
    candidate = None
    best_non_empty = 0
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        vals = _row_values(ws, r)
        non_empty = sum(1 for x in vals if x)
        line = " | ".join(vals)
        if non_empty > best_non_empty and (
            any(k in line for k in AGENT_KEYS) or
            TOTAL_KEY in line or TODAY_KEY in line or
            any(h in line for h in MONTH_HINTS)
        ):
            best_non_empty = non_empty
            candidate = r
    return candidate

def _headers_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    return { ("" if ws.cell(row=header_row, column=c).value is None else str(ws.cell(row=header_row, column=c).value).strip()): c
             for c in range(1, ws.max_column + 1) }

def _find_col_contains(hmap: Dict[str, int], must_contain: List[str]) -> Optional[int]:
    def norm(s: str) -> str:
        return s.replace("סכום של", "").replace("Σ", "").strip()
    for k, idx in hmap.items():
        kn = norm(k)
        if all(sub in kn for sub in must_contain):
            return idx
    return None

def _first_existing(hmap: Dict[str, int], keys: List[str]) -> Optional[int]:
    for k in keys:
        if k in hmap:
            return hmap[k]
    return None

def _is_total_label(val: object) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return s in {"Grand Total", "סה\"כ", "סהכ", "Total", "סך הכל", "סכום כולל"}

def _collect_month_cols(hmap: Dict[str, int]) -> List[int]:
    # כל כותרת שמכילה "חודש" או "טרם", ממויינת לפי אינדקס עמודה
    pairs = []
    for k, idx in hmap.items():
        if any(h in k for h in MONTH_HINTS):
            pairs.append((idx, k))
    pairs.sort(key=lambda t: t[0])
    return [idx for idx, _ in pairs]

def build_by_agent_sheet(
    processed_df: pd.DataFrame,   # לא בשימוש כרגע – שמור להרחבות
    output_path: str,
    pivot_sheet_name: str = SHEET_PIVOT_PRIVATE_DEFAULT,
    sheet_name: str = SHEET_BY_AGENT_DEFAULT,
    max_month_cols: int = 4,
    extra_pivot_sheet_name: Optional[str] = None,   # NEW: פיבוט תדמיתי
) -> Tuple[bool, str, int]:

    def _extract_from_pivot(ws: Worksheet) -> Tuple[List[Dict[str, object]], List[str]]:
        """מחזיר rows + רשימת כותרות חודשים כפי שמופיעות בפיבוט (עד max_month_cols)."""
        header_row = _find_header_row(ws)
        if not header_row:
            return [], []

        hmap = _headers_map(ws, header_row)

        agent_col_idx = _first_existing(hmap, AGENT_KEYS)
        if agent_col_idx is None:
            agent_col_idx = 1

        total_col_idx = _find_col_contains(hmap, [TOTAL_KEY])
        today_col_idx = _find_col_contains(hmap, [TODAY_KEY])

        # fallback אם עמודות סכומים לא זוהו
        if total_col_idx is None and today_col_idx is None:
            numeric_candidates = []
            for c in range(agent_col_idx + 1, ws.max_column + 1):
                val = ws.cell(row=header_row + 1, column=c).value
                if isinstance(val, (int, float)):
                    numeric_candidates.append(c)
                if len(numeric_candidates) >= 2:
                    break
            if numeric_candidates:
                if len(numeric_candidates) >= 1:
                    total_col_idx = numeric_candidates[0]
                if len(numeric_candidates) >= 2:
                    today_col_idx = numeric_candidates[1]

        # עמודות חודשים – עד 4
        month_cols_all = _collect_month_cols(hmap)
        month_cols = month_cols_all[:max_month_cols]

        # שמות כותרות החודשים כפי שהן בפיבוט
        month_headers_local = []
        for cidx in month_cols:
            for k, v in hmap.items():
                if v == cidx:
                    month_headers_local.append(k)
                    break

        rows_local: List[Dict[str, object]] = []
        r = header_row + 1
        while r <= ws.max_row:
            if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ws.max_column + 1)):
                break
            agent = ws.cell(row=r, column=agent_col_idx).value
            if agent and not _is_total_label(agent):
                def _get(cidx):
                    return ws.cell(row=r, column=cidx).value if cidx else None
                rec = {
                    "agent": agent,
                    "total": _get(total_col_idx),
                    "today": _get(today_col_idx),
                    "months": [ _get(c) for c in month_cols ],
                }
                rows_local.append(rec)
            r += 1
        return rows_local, month_headers_local

    # --- קריאה לפיבוטים בערכי חישוב (data_only) ---
    wb_vals = load_workbook(output_path, data_only=True)
    if pivot_sheet_name not in wb_vals.sheetnames:
        print(f"[לפי סוכן] לא נמצא גיליון פיבוט: {pivot_sheet_name}", flush=True)
        return False, sheet_name, 0

    rows_all: List[Dict[str, object]] = []
    month_headers_final: List[str] = []

    # פיבוט פרטי (חובה)
    pvt_private = wb_vals[pivot_sheet_name]
    rows_p, month_headers_p = _extract_from_pivot(pvt_private)
    for rec in rows_p:
        rec["channel"] = "שוק פרטי"
    rows_all.extend(rows_p)
    month_headers_final = list(month_headers_p)  # נתחיל מהראשון

    # פיבוט תדמיתי (אופציונלי)
    if extra_pivot_sheet_name and extra_pivot_sheet_name in wb_vals.sheetnames:
        pvt_ted = wb_vals[extra_pivot_sheet_name]
        rows_t, month_headers_t = _extract_from_pivot(pvt_ted)
        for rec in rows_t:
            rec["channel"] = "שוק תדמיתי"
        rows_all.extend(rows_t)
        # אם יש כותרות חודשים נוספות – נוסיף לסוף בסדר הופעה
        for h in month_headers_t:
            if h not in month_headers_final:
                month_headers_final.append(h)

    if not rows_all:
        print("[לפי סוכן] אין שורות נתונים לאחר סינון totals/ריקים.", flush=True)
        return False, sheet_name, 0

    # --- כתיבה עם נוסחאות/עיצוב ---
    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = [
        "סוכן","ערוץ","פיגור גביה חודש נוכחי","פיגור גביה חודש קודם","הפרש",
        TOTAL_KEY, TODAY_KEY
    ] + month_headers_final + ["סך פיגור"]
    ws.append(headers)

    def _num(v):
        try:
            return float(v) if v not in (None, "", "-") else 0.0
        except Exception:
            return 0.0

    start_row = 2
    months_count = len(month_headers_final)
    first_month_col_excel = 8  # H

    for i, rec in enumerate(rows_all):
        r_excel = start_row + i
        ws.cell(row=r_excel, column=1).value = rec["agent"]
        ws.cell(row=r_excel, column=2).value = rec.get("channel", "שוק פרטי")

        F = _num(rec.get("total"))
        G = _num(rec.get("today"))
        ws.cell(row=r_excel, column=6).value = F  # Total
        ws.cell(row=r_excel, column=7).value = G  # Today

        # חודשים (נרפד ל- months_count)
        months_vals = [ _num(x) for x in (rec.get("months") or []) ]
        if len(months_vals) < months_count:
            months_vals = months_vals + [0.0] * (months_count - len(months_vals))
        else:
            months_vals = months_vals[:months_count]

        for j, mv in enumerate(months_vals):
            ws.cell(row=r_excel, column=first_month_col_excel + j).value = mv

        # סך פיגור = SUM(H:עמודת החודש האחרונה)
        if months_count > 0:
            last_month_col_excel = first_month_col_excel + months_count - 1
            sum_col_excel = last_month_col_excel + 1
            ws.cell(row=r_excel, column=sum_col_excel).value = \
                f"=SUM({get_column_letter(first_month_col_excel)}{r_excel}:{get_column_letter(last_month_col_excel)}{r_excel})"
        else:
            sum_col_excel = 8
            ws.cell(row=r_excel, column=sum_col_excel).value = 0

        # C = IFERROR( (סך פיגור) / F , 0)
        ws.cell(row=r_excel, column=3).value = f"=IFERROR({get_column_letter(sum_col_excel)}{r_excel}/F{r_excel},0)"
        # D (פיגור קודם) נשאר ריק להזנה ידנית / העתקה מחודש קודם
        ws.cell(row=r_excel, column=4).value = None
        # E = C - D
        ws.cell(row=r_excel, column=5).value = f"=C{r_excel}-D{r_excel}"

    _style_header(ws)

    # פורמטים
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = PERCENT_FMT  # C
        ws.cell(row=r, column=4).number_format = PERCENT_FMT  # D
        ws.cell(row=r, column=5).number_format = PERCENT_FMT  # E
        for c in range(6, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = NUMBER_FMT

    _autosize(ws)
    wb.save(output_path)
    print(f"[לפי סוכן] נמצא {len(rows_all)} שורות | חודשי פיגור: {months_count}", flush=True)
    return True, sheet_name, len(rows_all)

