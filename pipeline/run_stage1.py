import argparse
import os
import shutil
import pandas as pd

#UI Design
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import PatternFill

from Logic.w10_load_and_unmerge import load_and_unmerge
from Logic.w15_detect_header import detect_header_and_frame
from Logic.w25_normalize_numeric_columns import normalize_numeric_columns
from Logic.w21_drop_specific_columns import drop_columns
from Logic.w55_remove_export_channel import remove_export_channel
from Logic.w50_remove_summary_rows import remove_summary_rows
from Logic.w52_normalize_agent_code import normalize_agent_code
from Logic.w27_drop_empty_rows import drop_empty_rows
from Logic.w28_filter_agent_code_required import filter_agent_code_required
from Logic.w30_add_sum_rows import append_sum_rows
from Logic.w40_finalize_save import save_processed
from Logic.w60_remove_other_rows import remove_other_rows  
from Logic.w71_manager_sheet_builder import build_manager_sheets
from Logic.w72_market_sheets import build_private_market_like_manager, build_tedmiti_full_columns
from Logic.w73_region_general_sheet import build_region_general_full_columns
from Logic.w75_pivot_sheets import build_pivot_private, build_pivot_tedmiti



from Logic.w77_fix_rafi_sheet import refine_rafi_sheet_rows
from Logic.w78_fix_private_sheet import refine_private_region_rows


# --- allow importing group_spec from project root ---
import sys, os
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)



def _argb(color_hex: str) -> str:
    """
    ממיר '#RRGGBB' או 'RRGGBB' ל- 'FFRRGGBB' (ARGB מלא).
    """
    s = color_hex.strip().lstrip("#")
    return ("FF" + s.upper()) if len(s) == 6 else s.upper()


def _color_manager_headers(
    xlsx_path: str,
    sheet_names: list[str],
    header_row: int = 1,
    col_H: str = "#BFEE90",
    col_I: str = "#90BFEE",
    cols_J_to_N: str = "#EEBF90",
):
    """
    צובע כותרות בלשוניות מנהלים:
    H = col_H,  I = col_I,  J..N = cols_J_to_N
     מתעלם מדפים/עמודות שאינם קיימים.
    """
    wb = load_workbook(xlsx_path)
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]

        def paint(idx: int, hex_color: str):
            if idx <= ws.max_column and idx >= 1 and hex_color:
                ws.cell(row=header_row, column=idx).fill = PatternFill(
                    fill_type="solid",
                    start_color=_argb(hex_color),
                    end_color=_argb(hex_color),
                )

        # H=8, I=9, J..N = 10..14
        paint(8, col_H)
        paint(9, col_I)
        last = min(ws.max_column, 14)
        for c in range(10, last + 1):
            paint(c, cols_J_to_N)

    wb.save(xlsx_path)
        
    
def _color_by_agent_headers(xlsx_path: str, sheet_name: str = "לפי סוכן",
                            header_row: int = 1,
                            col_H: str = "#BFEE90",
                            col_I: str = "#90BFEE",
                            cols_J_to_N: str = "#EEBF90"):
    """צובע כותרות ב'לפי סוכן': H, I, ו-J..N. בטוח להרצה חוזרת."""
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]

    def paint(idx: int, hex_color: str):
        if 1 <= idx <= ws.max_column and hex_color:
            ws.cell(row=header_row, column=idx).fill = PatternFill(
                fill_type="solid",
                start_color=_argb(hex_color),
                end_color=_argb(hex_color),
            )

    # H=8, I=9, J..N = 10..14 (רק אם קיימות)
    paint(8, col_H)
    paint(9, col_I)
    last = min(ws.max_column, 14)
    for c in range(10, last + 1):
        paint(c, cols_J_to_N)

    wb.save(xlsx_path)
    wb.close()
    

def _style_by_agent_columns(
    xlsx_path: str,
    sheet_name: str = "לפי סוכן",
    header_row: int = 1,
    hk_color: str = "#EEBF90",   # לכותרות H..K
    colL_color: str = "#E9FCE9"  # לכל עמודה L (כולל כותרת)
):
    """
    עיצוב לגליון 'לפי סוכן':
    - כותרות H..K בצבע hk_color
    - כל עמודה L (כולל הכותרת) בצבע colL_color
    """
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]

    # 1) כותרות H..K
    for col_idx in range(8, min(ws.max_column, 11) + 1):  # 8=H, 9=I, 10=J, 11=K
        ws.cell(row=header_row, column=col_idx).fill = PatternFill(
            fill_type="solid",
            start_color=_argb(hk_color),
            end_color=_argb(hk_color),
        )

    # 2) כל עמודה L (12) כולל הכותרת
    if ws.max_column >= 12:
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=12).fill = PatternFill(
                fill_type="solid",
                start_color=_argb(colL_color),
                end_color=_argb(colL_color),
            )

    wb.save(xlsx_path)
    wb.close()



def _outline_thick(xlsx_path: str, sheet_name: str = "לפי סוכן"):
    """מסגרת עבה (outline) לכל הטבלה + קווים דקים בפנים."""
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]

    thick = Side(style="medium")
    thin  = Side(style="thin")

    max_r, max_c = ws.max_row, ws.max_column
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            ws.cell(row=r, column=c).border = Border(
                left=thick if c == 1       else thin,
                right=thick if c == max_c  else thin,
                top=thick if r == 1        else thin,
                bottom=thick if r == max_r else thin,
            )

    wb.save(xlsx_path)
    wb.close()

def _shade_colA_and_group_borders(
    xlsx_path: str,
    sheet_name: str = "לפי סוכן",
    data_start_row: int = 2,     # הנתונים מתחילים בשורה 2 (כותרות בשורה 1)
    block_size: int = 5,         # כל 5 שורות קבוצה
    cycle_colors = ("#DDEFD6", "#FFF3A3", "#D6E8FF"),  # ירוק/צהוב/כחול עדין
):
    """
    צובע את עמודה A בקבוצות של 5 שורות במחזור צבעים,
    ומוסיף קו תחתון עבה בין כל קבוצה (כל 5 שורות).
    """
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    max_r, max_c = ws.max_row, ws.max_column
    if max_r < data_start_row:
        wb.close()
        return

    # הכנות למסגרות
    thin  = Side(style="thin")
    thick = Side(style="medium")

    # צביעה של עמודה A בקבוצות
    for r in range(data_start_row, max_r + 1):
        grp_idx = (r - data_start_row) // block_size
        color   = cycle_colors[grp_idx % len(cycle_colors)]
        ws.cell(row=r, column=1).fill = PatternFill(
            fill_type="solid",
            start_color=_argb(color),
            end_color=_argb(color),
        )

    # קו עבה בין כל 5 שורות (תחתון של שורת הגבול)
    for r in range(data_start_row + block_size - 1, max_r + 1, block_size):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            # משמרים קווים קיימים, רק מחליפים את התחתון לעבה
            cell.border = Border(
                left   = cell.border.left   or thin,
                right  = cell.border.right  or thin,
                top    = cell.border.top    or thin,
                bottom = thick,
            )

    wb.save(xlsx_path)
    wb.close()
    

FRONT_CANDIDATES = [
    "מנהל סחר","מנהל אזור","מנהל איזור",
    "סוכן","ערוץ","שיטת תשלום לקוח משלם","קוד לקוח קצה","לקוח קצה","קוד סוכן"
]
SUM_HEADER = 'סה"כ סכום יתרת חוב'


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="נתיב לקובץ המקור (xlsx)")
    parser.add_argument("--output-dir", required=True, help="תיקיית פלט לשמירת הקובץ המעובד")
    parser.add_argument("--sheet-name", default="sheet1", help="שם הגיליון המקורי (ברירת מחדל: sheet1)")
    parser.add_argument("--sum-header", default=SUM_HEADER, help="שם העמודה בה נחשב סכום בסוף")
    parser.add_argument("--drop-empty", action="store_true", help="מחיקת שורות ריקות/חסרות מזהי מפתח")
    parser.add_argument("--keep-other", action="store_true", help="אל תנקה 'אחר'/'אחר אחר' במנהלי סחר/אזור")
    parser.add_argument("--keep-temp", action="store_true", help="השארת קבצים זמניים לבדיקה")
    parser.add_argument("--with-nov-dec", dest="with_nov_dec", action="store_true",
                        help="[תאימות לאחור] לא בשימוש במצב דינמי (I..M אחרי H)")

    # דוחות נגזרים
    parser.add_argument("--split-by-manager", action="store_true",
                        help="יצירת לשוניות לכל 'מנהל סחר' מתוך גיליון 'מעובד' (J..M דינמי + 'טור עזר').")
    parser.add_argument("--market-private", action="store_true",
                        help="יצירת גיליון 'שוק פרטי' (מנהל סחר=רפי מור יוסף-סחר, ערוץ=שוק פרטי) במבנה כמו מנהלי סחר")
    parser.add_argument("--market-tedmiti", action="store_true",
                        help="יצירת גיליון 'שוק תדמיתי' (מנהל סחר=עמי חכמון) עם כל העמודות מה'מעובד'")
    parser.add_argument("--region-general", action="store_true",
                        help="יצירת גיליון 'מנהל אזור כללי' עם כל העמודות מ'מעובד'")
    parser.add_argument("--pivot-private", action="store_true",
                        help="יצירת גיליון 'פיבוט פרטי' (מנהל סחר=רפי מור יוסף-סחר, ערוץ=שוק פרטי)")
    parser.add_argument("--pivot-tedmiti", action="store_true",
                        help="יצירת גיליון 'פיבוט תדמיתי' (מנהל סחר=עמי חכמון)")
    
    parser.add_argument("--by-agent", action="store_true",
                    help="יצירת גיליון 'לפי סוכן' מתוך 'פיבוט פרטי'")


    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    temp_dir = os.path.join(args.output_dir, "_temp")
    os.makedirs(temp_dir, exist_ok=True)

    # חישוב צעדים עד שמירת 'מעובד' (הדוחות הנגזרים אינם נספרים בלוג זה)
    total_steps = 12 + (0 if args.keep_other else 1) + (1 if args.drop_empty else 0)
    step = 1

    print(f"[{step}/{total_steps}] ביטול מיזוגים ושמירת קובץ זמני...", flush=True); step += 1
    tmp_path = load_and_unmerge(args.input, sheet_hint=args.sheet_name, temp_out_dir=temp_dir)

    print(f"[{step}/{total_steps}] איתור שורת כותרות ובניית DataFrame...", flush=True); step += 1
    df_all = detect_header_and_frame(tmp_path, args.sheet_name)

    if args.with_nov_dec:
        print("הערה: --with-nov-dec מתעלמים ממנו במצב דינמי (I..M אחרי H).", flush=True)

    print(f"[{step}/{total_steps}] בחירת וסידור עמודות (דינמי I..M אחרי '{args.sum_header}')...", flush=True); step += 1
    # חזית קבועה
    front = [c for c in FRONT_CANDIDATES if c in df_all.columns]
    # 5 הכותרות שאחרי sum-header במקור
    all_cols = list(df_all.columns)
    dyn5 = []
    if args.sum_header in all_cols:
        h_idx = all_cols.index(args.sum_header)
        for j in range(1, 6):
            if h_idx + j < len(all_cols):
                dyn5.append(all_cols[h_idx + j])
    dyn5 = [c for c in dyn5 if c in df_all.columns]

    # סדר סופי: עד "קוד סוכן" -> ואז sum-header + dyn5
    final_cols = list(front)
    if args.sum_header in df_all.columns:
        if "קוד סוכן" in final_cols:
            insert_at = final_cols.index("קוד סוכן") + 1
            block = [args.sum_header] + dyn5  # בסדר המקורי
            for item in reversed(block):
                if item in df_all.columns and item not in final_cols:
                    final_cols.insert(insert_at, item)
        else:
            if args.sum_header not in final_cols:
                final_cols.append(args.sum_header)
            for c in dyn5:
                if c not in final_cols:
                    final_cols.append(c)

    df_proc = df_all.copy()
    for c in final_cols:
        if c not in df_proc.columns:
            df_proc[c] = None
    df_proc = df_proc[final_cols]

    print(f"[{step}/{total_steps}] נרמול ערכים מספריים ('{args.sum_header}' ועוד 5 הדינמיות)...", flush=True); step += 1
    amount_headers = [h for h in [args.sum_header] + dyn5 if h in df_proc.columns]
    df_proc = normalize_numeric_columns(df_proc, amount_headers)

    print(f"[{step}/{total_steps}] מחיקת עמודות לא נדרשות...", flush=True); step += 1
    df_proc, dropped_cols = drop_columns(df_proc, columns=("שייייטת תשלום")) #("שיטת תשלום לקוח משלם","שיטת תשלום") -> מחקתי שיטת תשלום לקוח משלם כי הלקוח החליט לשחזר עמודה
    if dropped_cols:
        print(f"    הוסרו עמודות: {dropped_cols}", flush=True)

    print(f"[{step}/{total_steps}] מחיקת ערוץ 'ייצוא'...", flush=True); step += 1
    df_proc, _ = remove_export_channel(df_proc, col="ערוץ")

    print(f"[{step}/{total_steps}] מחיקת שורות סיכום לפי 'קוד סוכן'...", flush=True); step += 1
    df_proc, _ = remove_summary_rows(df_proc, col="קוד סוכן")

    print(f"[{step}/{total_steps}] נירמול שמות 'סוכן' (יעל כץ...)...", flush=True); step += 1
    if "סוכן" in df_proc.columns:
        df_proc["סוכן"] = df_proc["סוכן"].astype(str).str.strip().replace({
            "יעל כץ מלונות": "יעל כץ",
            "יעל כץ תדמיתי": "יעל כץ",
        })
        
     # סינון שורות לא רצויות בעמודת 'סוכן'
    if "סוכן" in df_proc.columns:
        bad_agents = {"חובות מסופקים", "לקוחות שוק קמעונאי"}
        s_clean = df_proc["סוכן"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
        removed_mask = s_clean.isin(bad_agents)
        n_removed = int(removed_mask.sum())
        if n_removed:
            df_proc = df_proc[~removed_mask]
            print(f"    הוסרו {n_removed} שורות ('סוכן' בעייתי).", flush=True)
   

    print(f"[{step}/{total_steps}] נירמול 'קוד סוכן' לספרות בלבד...", flush=True); step += 1
    df_proc, _ = normalize_agent_code(df_proc, col="קוד סוכן")

    if not args.keep_other:
        print(f"[{step}/{total_steps}] מחיקת 'אחר'/'אחר אחר' ב'מנהל סחר/אזור'...", flush=True); step += 1
        df_proc, _ = remove_other_rows(df_proc, cols=["מנהל סחר","מנהל אזור","מנהל איזור"])

    if args.drop_empty:
        print(f"[{step}/{total_steps}] מחיקת שורות ריקות/חסרות מזהים...", flush=True); step += 1
        df_proc, rem_full, rem_req = drop_empty_rows(df_proc, required_any=["קוד סוכן","קוד לקוח קצה","לקוח קצה"])
        print(f"    הוסרו {rem_full} ריקות ו-{rem_req} ללא מזהים.", flush=True)

    # *** סינון סופי לפני סכום: 'קוד סוכן' חייב להיות מספרי ***
    print(f"[{step}/{total_steps}] סינון סופי: 'קוד סוכן' – רק ספרות ללא ריקים...", flush=True); step += 1
    df_proc, removed_non_numeric = filter_agent_code_required(df_proc, col="קוד סוכן")
    print(f"    הוסרו {removed_non_numeric} שורות ללא 'קוד סוכן' מספרי.", flush=True)

    print(f"[{step}/{total_steps}] הוספת שורות סכום בסוף '{args.sum_header}'...", flush=True); step += 1
    df_proc = append_sum_rows(df_proc, args.sum_header)

    print(f"[{step}/{total_steps}] שמירה בשם עם חותמת זמן וגיליון 'מעובד'...", flush=True); step += 1
    out_path = save_processed(df_proc, input_path=args.input, output_dir=args.output_dir, sheet_name="מעובד")

    # ===== דוחות נגזרים מתוך 'מעובד' =====
    # לעבודה על דוחות: מסירים שורות "סכום" טקסטואליות
    df_for_reports = df_proc.copy()
    if args.sum_header in df_for_reports.columns:
        df_for_reports = df_for_reports[
            ~(df_for_reports[args.sum_header].astype(str).str.strip() == "סכום")
        ]

    print("\n[דוחות נגזרים] בנייה לפי דגלים...", flush=True)


    # 1) לשוניות מנהלי סחר
    if args.split_by_manager:
        print("• בניית לשוניות מנהלי סחר...", flush=True)
        try:
            # בסיס הנתונים ללשוניות מנהלים
            df_mgr = df_for_reports.copy()
            if "מנהל סחר" in df_mgr.columns:
                # 1) נפטרים מ-NaN אמיתיים
                df_mgr = df_mgr[df_mgr["מנהל סחר"].notna()].copy()
                # 2) נרמול רווחים וקצוות
                df_mgr["מנהל סחר"] = (
                    df_mgr["מנהל סחר"].astype(str)
                    .str.replace(r"\s+", " ", regex=True)
                    .str.strip()
                )
                # 3) מסירים ערכי־דמה שעלולים לייצר לשונית בשם None/nan
                df_mgr = df_mgr[~df_mgr["מנהל סחר"].str.fullmatch(r"(?i)none|nan|null|", na=False)]
                # 4) מדלגים על "עמי חכמון" כדי שלא תיווצר לו לשונית
                df_mgr = df_mgr[df_mgr["מנהל סחר"] != "עמי חכמון"]

            n_created, names = build_manager_sheets(
                df_mgr, out_path,
                managers_col="מנהל סחר",
                max_month_cols_after_today=4
            )
            print(f"    נוצרו {n_created} גיליונות מנהלים: {', '.join(names)}", flush=True)
            
                    # צביעת כותרות H/I/J..N בלשוניות המנהלים בלבד
            try:
                _color_manager_headers(
                    out_path,
                    names,
                    header_row=1,
                    col_H="#BFEE90",
                    col_I="#90BFEE",
                    cols_J_to_N="#EEBF90",
                )
                print("    עודכנו צבעי כותרות בלשוניות המנהלים (H/I/J..N).", flush=True)
            except Exception as e:
                print(f"    [אזהרה] כשל בצביעת כותרות: {e}", flush=True)


        # סינון/נרמול לשונית(ות) רפי – רק שורות שבהן 'מנהל אזור/איזור' הוא רפי,
        # ואח"כ עדכון הטקסט ל"רפי מור יוסף- סחר"
            ok_refine, touched, del_counts = refine_rafi_sheet_rows(
                out_path,
                target_base="רפי מור יוסף",
                display_text="רפי מור יוסף- סחר"
            )
            if ok_refine:
                total_del = sum(del_counts)
                print(f"    [רפי] סוננו {total_del} שורות בלשוניות: {', '.join(touched)}", flush=True)
            else:
                print("    [רפי] לא נמצאו שורות לסינון/שינוי.", flush=True)

        except Exception as e:
            print(f"שגיאה בבניית לשוניות מנהלים/רפי: {e}", flush=True)


    # 2) שוק פרטי
    if args.market_private:
        print("• בניית גיליון 'שוק פרטי'...", flush=True)
        try:
            ok, name = build_private_market_like_manager(
                df_for_reports, out_path,
                manager_name="רפי מור יוסף- סחר",  # השאר כמו אצלך
                channel_value="שוק פרטי",
                max_month_cols_after_today=4,
                sheet_name="שוק פרטי"
            )
            if ok:
                # פוסט-סינון: מחיקת שורות שבהן 'מנהל אזור/איזור' = 'רפי מור יוסף'
                did, deleted = refine_private_region_rows(
                    out_path,
                    sheet_name=name,                 # משתמשים בשם שחזר מה-builder
                    forbidden_substr="רפי מור יוסף- סחר"
                )
                if did:
                    print(f"    נבנה: {name} | הוסרו {deleted} שורות ('מנהל אזור'='רפי מור יוסף- סחר')", flush=True)
                else:
                    print(f"    נבנה: {name} | אין שורות למחיקה בעמודת 'מנהל אזור'", flush=True)
            else:
                print("    לא נבנה (אין נתונים)", flush=True)
        except Exception as e:
            print(f"שגיאה בבניית גיליון 'שוק פרטי': {e}", flush=True)


    # 3) שוק תדמיתי
    if args.market_tedmiti:
        print("• בניית גיליון 'שוק תדמיתי'...", flush=True)
        try:
            ok, name = build_tedmiti_full_columns(
                df_for_reports, out_path,
                manager_name="עמי חכמון",
                sheet_name="שוק תדמיתי"
            )
            print(f"    נבנה: {name}" if ok else "    לא נבנה (אין נתונים)", flush=True)
        except Exception as e:
            print(f"שגיאה בבניית גיליון 'שוק תדמיתי': {e}", flush=True)

    # 4) מנהל אזור כללי
    if args.region_general:
        print("• בניית גיליון 'מנהל אזור כללי'...", flush=True)
        try:
            ok, name = build_region_general_full_columns(
                df_for_reports, out_path, sheet_name="מנהל אזור כללי"
            )
            print(f"    נבנה: {name}" if ok else "    לא נבנה", flush=True)
        except Exception as e:
            print(f"שגיאה בבניית 'מנהל אזור כללי': {e}", flush=True)

    # 5) פיבוט פרטי
    if args.pivot_private:
        print("• בניית גיליון 'פיבוט פרטי'...", flush=True)
        try:
            ok, name = build_pivot_private(
                df_for_reports, out_path,
                manager_name="רפי מור יוסף-סחר",
                channel_value="שוק פרטי",
                sheet_name="פיבוט פרטי",
                max_month_cols_after_today=4
            )
            print(f"    נבנה: {name}" if ok else "    לא נבנה (אין נתונים)", flush=True)
        except Exception as e:
            print(f"שגיאה בבניית גיליון 'פיבוט פרטי': {e}", flush=True)

    # 6) פיבוט תדמיתי
    if args.pivot_tedmiti:
        print("• בניית גיליון 'פיבוט תדמיתי'...", flush=True)
        try:
            ok, name = build_pivot_tedmiti(
                df_for_reports, out_path,
                manager_name="עמי חכמון",
                sheet_name="פיבוט תדמיתי",
                max_month_cols_after_today=4
            )
            print(f"    נבנה: {name}" if ok else "    לא נבנה (אין נתונים)", flush=True)
        except Exception as e:
            print(f"שגיאה בבניית גיליון 'פיבוט תדמיתי': {e}", flush=True)
            


    # 7) לפי סוכן (w90)
    if args.by_agent:
        print("• בניית גיליון 'לפי סוכן' (w90)...", flush=True)
        ok = False
        try:
            from Logic.w90_agent import build_by_agent_sheet_w90
            ok, name, nrows = build_by_agent_sheet_w90(
                out_path,
                private_pivot="פיבוט פרטי",
                tedmiti_pivot="פיבוט תדמיתי",
                sheet_name="לפי סוכן",
            )
            print(f"    נבנה: {name} (שורות: {nrows})" if ok else "    לא נבנה (אין נתונים/לא נמצא פיבוט)", flush=True)
        except Exception as e:
            print(f"שגיאה בבניית 'לפי סוכן' (w90): {e}", flush=True)

        if ok:
        # (2) שורות ריקות לקבוצות
            try:
                from Logic.w90_agent import ensure_group_blank_rows_w90
                ensure_group_blank_rows_w90(out_path, sheet_name="לפי סוכן")
            except Exception as ge:
                print(f"    [אזהרה] הוספת שורות ריקות נכשלה: {ge}", flush=True)

        # (3) ריתוך נוסחאות הסיכום אחרי ההחדרה
            try:
                from Logic.w90_agent import rebind_all_sum_rows_w90
                rebind_all_sum_rows_w90(out_path, sheet_name="לפי סוכן")
            except Exception as ge:
                print(f"    [אזהרה] תיקון נוסחאות סיכום נכשל: {ge}", flush=True)

        # (4) רשתות ארציות
            try:
                from Logic.w90_agent import link_national_from_manager_sheets_w90
                link_national_from_manager_sheets_w90(out_path, sheet_name="לפי סוכן")
                print("    רשתות ארציות מולאו מנתוני גיליונות המנהלים.", flush=True)
            except Exception as ge:
                print(f"    [אזהרה] רשתות ארציות לא מולאו: {ge}", flush=True)

        # (5) L/C/E
            try:
                from Logic.w90_agent import ensure_pigor_sum_and_pct_w90
                ensure_pigor_sum_and_pct_w90(out_path, sheet_name="לפי סוכן")
                print("    עודכן: 'סך פיגור' (L) + אחוזים ב-C/E בכל השורות.", flush=True)
            except Exception as ge:
                print(f"    [אזהרה] חישוב 'סך פיגור' ואחוזים נכשל: {ge}", flush=True)

        # (6) פריסת עמודות למניעת #####
            try:
                from Logic.w90_agent import set_column_layout_w90
                set_column_layout_w90(out_path, sheet_name="לפי סוכן")
                print("    עיצוב תצוגה: רוחבי עמודות + RTL + Freeze + shrink-to-fit (למניעת #####).", flush=True)
            except Exception as ge:
                print(f"    [אזהרה] עיצוב תצוגה לא הושלם: {ge}", flush=True)

        # (7) צבעי קבוצות/קו עבה/Bold
            try:
                from Logic.w90_agent import style_groups_colA_only_w90
                style_groups_colA_only_w90(out_path, sheet_name="לפי סוכן")
                print("    עיצוב: עמודה A בקבוצות + Bold לשורות המבוקשות + קווים.", flush=True)
            except Exception as ge:
                print(f"    [אזהרה] עיצוב-סיום לא הושלם: {ge}", flush=True)
                
                
        #remove bold line between lines 21 and 22
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Border

                wb = load_workbook(out_path)
                if "לפי סוכן" in wb.sheetnames:
                    ws = wb["לפי סוכן"]

                    # מצא את השורה של 'סה"כ מוקד'
                    row_sum = None
                    for r in range(2, ws.max_row + 1):
                        a = str(ws.cell(row=r, column=1).value or "").strip()
                        if a == 'סה"כ מוקד':
                            row_sum = r
                            break

                    # אם השורה הבאה היא 'אריק יחזקאל' – ננקה את ה-bottom border בשורת 'סה"כ מוקד'
                    if row_sum and row_sum + 1 <= ws.max_row:
                        next_a = str(ws.cell(row=row_sum + 1, column=1).value or "").strip()
                        if next_a == "אריק יחזקאל":
                            for c in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_sum, column=c)
                                b = cell.border
                                cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=None)

                    wb.save(out_path)
            except Exception as ge:
                print(f'    [אזהרה] ניקוי קו עבה אחרי "סה\\"כ מוקד" נכשל: {ge}', flush=True)

        # (8) בלי קו עבה מעל 'סה\"כ רשתות ארציות'
            try:
                from Logic.w90_agent import remove_thick_above_national_total
                remove_thick_above_national_total(out_path, sheet_name="לפי סוכן")
                print("    עודכן: הוסר הקו העבה מעל 'סה\"כ רשתות ארציות'.", flush=True)
            except Exception as ge:
                print(f"    [אזהרה] ניקוי קו עבה מעל רשתות ארציות נכשל: {ge}", flush=True)
                
            

    print("\nהפקה הושלמה בהצלחה:", out_path)
    
    if not args.keep_temp:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass


if __name__ == "__main__":
    main()
    
    
    
    
    
    
    
